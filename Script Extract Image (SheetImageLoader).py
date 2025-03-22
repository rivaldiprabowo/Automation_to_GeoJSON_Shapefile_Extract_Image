# %% [markdown]
# # Application to Extract Image from Excel

# %% [markdown]
# ## 1. Import Library

# %%
import os
import pandas as pd
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image
import io
import re
import gc
from concurrent.futures import ThreadPoolExecutor
import time
from io import BytesIO
from openpyxl.utils import column_index_from_string, get_column_letter
from lxml import etree
import zipfile
import shutil
import xml.etree.ElementTree as ET
from PIL import Image as PILImage
from collections import defaultdict

# %% [markdown]
# ## 2. Application to Extract Images from Excel

# %% [markdown]
# ### 2.1. Function Codes

# %%
import re
import os
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter

def extract_images_from_excel(file_path, output_folder): #Extract images from columns containing specific keywords in all sheets and save them.

    # Keywords to look for in column names
    KEYWORDS = ["DOKUMENTASI", "RAMBU", "RPPJ"]
    
    # Extract filename from path for folder creation
    file_name_clean = re.search(r'([^\\]+)\.xlsx$', file_path)
    if file_name_clean:
        file_name_clean = file_name_clean.group(1)
    else:
        file_name_clean = os.path.basename(file_path).replace('.xlsx', '')
    
    # Create subdirectories for each category
    dokumentasi_folder = os.path.join(output_folder, "Dokumentasi", file_name_clean)  # Changed - nested under Excel name
    rambu_folder = os.path.join(output_folder, "Rambu")
    rppj_folder = os.path.join(output_folder, "RPPJ")
    
    for folder in [dokumentasi_folder, rambu_folder, rppj_folder]:
        os.makedirs(folder, exist_ok=True)
    
    try:
        # Process each sheet independently to prevent file handle issues
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames.copy()
        wb.close()
        
        successful_sheets = 0
        total_images_saved = 0
        
        # Process sheets in the order they appear in the workbook
        for sheet_idx, sheet_name in enumerate(sheet_names, 1):
            try:
                print(f"Processing sheet {sheet_idx}/{len(sheet_names)}: {sheet_name}")
                
                # Create a safe sheet name for filenames
                safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
                
                # Open a fresh workbook instance for each sheet
                wb = load_workbook(file_path, data_only=True)
                ws = wb[sheet_name]
                
                # Create a fresh image loader for this sheet
                image_loader = SheetImageLoader(ws)
                
                # Step 1: Extract merged column headers from rows 1-5
                column_names = {}
                
                # Get values for rows 1-5 for each column
                for col in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col)
                    header_values = []
                    
                    for row in range(1, 6):  # Rows 1-5
                        cell_value = ws[f"{col_letter}{row}"].value
                        if cell_value:
                            header_values.append(str(cell_value).strip())
                    
                    # Combine the header parts into one name
                    if header_values:
                        column_names[col] = " ".join(header_values)
                    else:
                        column_names[col] = f"Column_{col_letter}"
                
                # Step 2: Find columns containing our keywords
                dokumentasi_columns = {}
                rambu_columns = {}
                rppj_columns = {}
                nama_rambu_column = None
                jenis_tiang_column = None
                
                # First, find all our target columns
                for col, name in column_names.items():
                    if "DOKUMENTASI" in name.upper():
                        dokumentasi_columns[col] = name
                    elif "RAMBU" in name.upper() and "NAMA RAMBU" not in name.upper():
                        rambu_columns[col] = name
                    elif "RPPJ" in name.upper():
                        rppj_columns[col] = name
                    
                    # Find the specific column for "Nama Rambu"
                    if "NAMA RAMBU" in name.upper():
                        nama_rambu_column = col
                        print(f"Found 'Nama Rambu' column: {name} (Column {get_column_letter(col)})")
                    
                    # Find the specific column for "Jenis Tiang"
                    elif "JENIS TIANG" in name.upper():
                        jenis_tiang_column = col
                        print(f"Found 'Jenis Tiang' column: {name} (Column {get_column_letter(col)})")
                
                # Track images processed for each category
                images_by_category = {
                    "dokumentasi": 0,
                    "rambu": 0,
                    "rppj": 0
                }
                
                # Process each category separately
                # 1. Process DOKUMENTASI columns
                if dokumentasi_columns:
                    processed = process_image_columns(
                        ws, image_loader, dokumentasi_columns, dokumentasi_folder, 
                        file_name_clean, safe_sheet_name, "dokumentasi", None, None
                    )
                    images_by_category["dokumentasi"] += processed
                
                # 2. Process RAMBU columns (custom naming based on "Nama Rambu" column)
                if rambu_columns:
                    if nama_rambu_column is None:
                        print("⚠️ Warning: 'Nama Rambu' column not found. Using default naming for Rambu images.")
                    
                    processed = process_image_columns(
                        ws, image_loader, rambu_columns, rambu_folder, 
                        file_name_clean, safe_sheet_name, "rambu", nama_rambu_column, None
                    )
                    images_by_category["rambu"] += processed
                
                # 3. Process RPPJ columns (custom naming based on "Jenis Tiang" column)
                if rppj_columns:
                    if jenis_tiang_column is None:
                        print("⚠️ Warning: 'Jenis Tiang' column not found. Using default naming for RPPJ images.")
                    
                    processed = process_image_columns(
                        ws, image_loader, rppj_columns, rppj_folder, 
                        file_name_clean, safe_sheet_name, "rppj", None, jenis_tiang_column
                    )
                    images_by_category["rppj"] += processed
                
                total_images_saved += sum(images_by_category.values())
                
                print(f"✅ Completed sheet '{sheet_name}':")
                print(f"  - Dokumentasi: {images_by_category['dokumentasi']} images")
                print(f"  - Rambu: {images_by_category['rambu']} images")
                print(f"  - RPPJ: {images_by_category['rppj']} images")
                
                if sum(images_by_category.values()) > 0:
                    successful_sheets += 1
                
            except Exception as e:
                print(f"❌ Error processing sheet '{sheet_name}' in file '{file_name_clean}': {str(e)}")
            
            finally:
                # Always close the workbook after processing each sheet
                if 'wb' in locals() and wb is not None:
                    try:
                        wb.close()
                    except:
                        pass
        
        print(f"✅ Completed processing file: {file_name_clean}")
        print(f"  - {successful_sheets}/{len(sheet_names)} sheets processed")
        print(f"  - {total_images_saved} total images extracted")
        return True
        
    except Exception as e:
        # Handle any errors in the outer scope
        print(f"❌ Error processing file '{file_path}': {str(e)}")
        return False

def process_image_columns(ws, image_loader, target_columns, output_folder, file_name_clean, 
                          safe_sheet_name, category, nama_rambu_column, jenis_tiang_column): #Process images in specified columns with custom naming logic.
    
    # Track existing image names (to avoid duplicates for Rambu)
    existing_images = {}
    
    # Find image cells in target columns
    image_cells = {}
    for col in target_columns:
        col_letter = get_column_letter(col)
        column_name = target_columns[col]
        
        # Make column name safe for filename
        safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
        
        # Scan rows starting from row 6
        for row in range(6, ws.max_row + 1):
            cell_address = f"{col_letter}{row}"
            if image_loader.image_in(cell_address):
                image_cells[(row, col)] = {
                    'cell_address': cell_address,
                    'column_name': safe_column_name,
                    'column_letter': col_letter,
                    'row_number': row
                }
    
    if not image_cells:
        print(f"⚠️ No images found in {category.upper()} columns, skipping...")
        return 0
    
    print(f"Found {len(image_cells)} images in {category.upper()} columns")
    
    # Process images in column-first, row-second order
    successful_images = 0
    image_cells_by_column = {}
    
    # Group by column
    for (row, col), cell_info in image_cells.items():
        if col not in image_cells_by_column:
            image_cells_by_column[col] = []
        image_cells_by_column[col].append((row, cell_info))
    
    # Process each column
    for col in sorted(image_cells_by_column.keys()):
        column_letter = get_column_letter(col)
        column_name = target_columns[col]
        safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
        
        print(f"Processing {len(image_cells_by_column[col])} images in column '{column_name}'")
        
        # Process rows in order
        for row, cell_info in sorted(image_cells_by_column[col], key=lambda x: x[0]):
            cell_address = cell_info['cell_address']
            try:
                # Get the image
                img = image_loader.get(cell_address)
                
                # Generate filename based on category
                if category == "rambu" and nama_rambu_column is not None:
                    # For RAMBU: Use "Nama Rambu" column value as filename
                    nama_rambu_value = ws[f"{get_column_letter(nama_rambu_column)}{row}"].value
                    
                    if nama_rambu_value and str(nama_rambu_value).strip():
                        # Use the actual value from "Nama Rambu" column
                        safe_nama_rambu = re.sub(r'[\\/*?:"<>|]', "_", str(nama_rambu_value).strip())
                        img_filename = f"{safe_nama_rambu}.png"
                        
                        # Check if this name already exists (to avoid duplicates)
                        if img_filename in existing_images:
                            print(f"  ⚠️ Duplicate 'Nama Rambu' found: {safe_nama_rambu} - Replacing existing image")
                        
                        existing_images[img_filename] = True
                    else:
                        # Fallback to default naming if no nama_rambu value
                        print(f"  ⚠️ No 'Nama Rambu' value found for row {row}, using default naming")
                        row_identifier = f"Row{row}"
                        safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                        img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                
                elif category == "rppj" and jenis_tiang_column is not None:
                    # For RPPJ: Use "Jenis Tiang" column value in filename
                    jenis_tiang_value = ws[f"{get_column_letter(jenis_tiang_column)}{row}"].value
                    safe_jenis_tiang = "Unknown"
                    if jenis_tiang_value and str(jenis_tiang_value).strip():
                        safe_jenis_tiang = re.sub(r'[\\/*?:"<>|]', "_", str(jenis_tiang_value).strip())
                    
                    # Get row identifier - ALWAYS use just the row number without column A value
                    row_identifier = f"Row{row}"
                    
                    safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                    img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_jenis_tiang}_{safe_row_identifier}.png"
                
                else:
                    # Default naming for DOKUMENTASI (and fallback for others)
                    # ALWAYS use just the row number without column A value
                    row_identifier = f"Row{row}"
                    
                    safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                    img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                
                # Save the image
                img_path = os.path.join(output_folder, img_filename)
                with io.BytesIO() as img_buffer:
                    img.save(img_buffer, format="PNG")
                    img_buffer.seek(0)
                    with open(img_path, 'wb') as f:
                        f.write(img_buffer.read())
                
                successful_images += 1
                print(f"  ✅ Saved: {img_filename}")
            except Exception as e:
                print(f"  ❌ Error saving image at {cell_address}: {str(e)}")
    
    return successful_images

def process_single_excel_file(file_path, export_folder): #Process a single Excel file and extract images from it.
    # Create "Extract Images" folder within the export directory
    output_folder = os.path.join(export_folder, "Extract Images")
    os.makedirs(output_folder, exist_ok=True)
    
    # Process the file
    file_name = os.path.basename(file_path)
    print(f"\n📊 Processing file: {file_name}")
    
    result = extract_images_from_excel(file_path, output_folder)
    
    # Print summary
    print("\n" + "="*50)
    print("📈 PROCESSING SUMMARY")
    print("="*50)
    print(f"File: {file_name}")
    
    if result:
        print(f"Status: Successfully processed ✅")
    else:
        print(f"Status: Failed to process ❌")
    
    print(f"Images saved to: {output_folder}")
    print("\n🎉 Processing completed!")
    
    return result

def process_excel_folder_images(folder_path, export_folder): #Process all Excel files in a folder and extract images from them.

    # Create "Extract Images" folder within the export directory
    output_folder = os.path.join(export_folder, "Extract Images")
    os.makedirs(output_folder, exist_ok=True)
    
    # Track statistics
    total_files = 0
    successful_files = 0
    failed_files = []
    
    # Get all Excel files in the folder - with full paths
    excel_files = []
    for file in os.listdir(folder_path):
        if file.endswith(('.xlsx', '.xlsm')):
            excel_files.append(os.path.join(folder_path, file))
    
    if not excel_files:
        print("⚠️ No Excel files found in the specified folder.")
        return
    
    total_files = len(excel_files)  # Fixed variable assignment
    
    # Process each Excel file
    for i, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print(f"\n📊 Processing file {i}/{total_files}: {file_name}")
        
        if extract_images_from_excel(file_path, output_folder):
            successful_files += 1
        else:
            failed_files.append(file_name)
    
    # Print summary
    print("\n" + "="*50)
    print("📈 PROCESSING SUMMARY")
    print("="*50)
    print(f"Total files: {total_files}")
    print(f"Successfully processed: {successful_files}")
    print(f"Failed to process: {len(failed_files)}")
    print(f"Images saved to: {output_folder}")
    
    if failed_files:
        print("\nFiles that could not be processed:")
        for file in failed_files:
            print(f"- {file}")
    
    print("\n🎉 All Excel files processing completed!")

# %% [markdown]
# ### 2.2. Run Function

# %%
excel_folder = r"C:\Users\kanzi\Documents\Part Time Job\Data Hasil Survey\01. Cileungsi - Cibeet.xlsx"  # Path to Excel files
export_folder = r"C:\Users\kanzi\Documents\Part Time Job\Hasil Export1"  # Path for export results
process_single_excel_file(excel_folder, export_folder) # Run the function with your paths

# %%
excel_folder = r"C:\Users\kanzi\Documents\Part Time Job\Data Hasil Survey"  # Path to Excel files
export_folder = r"C:\Users\kanzi\Documents\Part Time Job\Hasil Export1"  # Path for export results
process_excel_folder_images(excel_folder, export_folder) # Run the function with your paths


