# %% [markdown]
# # Automation Excel to GeoJSON and Shapefile

# %% [markdown]
# ## 1. Import Library

# %%
import pandas as pd
from openpyxl import load_workbook
import geopandas as gpd
from shapely.geometry import Point, LineString
import json
import glob
import os
from pathlib import Path

# %% [markdown]
# ## 2. Application to Export Excel into GeoJson

# %% [markdown]
# ### 2.1. Function Codes

# %%
def unique_column_names(columns): # Ensure column names are unique by appending suffix.
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def clean_column_names(columns): # Standardize column names by capitalizing each word properly.
    cleaned_columns = []
    seen = {}
    for col in columns:
        col = str(col).strip()
        
        # Remove columns with "Rekap" in the name
        if "rekap" in col.lower():
            continue
        
        # Fix for duplicated words (like "No No", "Detail Lokasi Detail Lokasi")
        words = col.split()
        if len(words) >= 2:
            # Check for repeated word patterns
            half_length = len(words) // 2
            if words[:half_length] == words[half_length:] and len(words) % 2 == 0:
                # If the first half matches the second half, only use the first half
                col = " ".join(words[:half_length])
        
        # Capitalize words
        col = " ".join(word.capitalize() for word in col.split())
        
        # Handle duplicates
        if col in seen:
            seen[col] += 1
            col = f"{col} {seen[col]}"
        else:
            seen[col] = 0
        
        cleaned_columns.append(col)
    return cleaned_columns

def fix_coordinates(row, lat_col, lon_col): #Fix latitude and longitude values that may be in the wrong format.
    lat, lon = row[lat_col], row[lon_col]
    if pd.notna(lat) and abs(lat) > 90:
        lat /= 1_000_000
    if pd.notna(lon) and abs(lon) > 180:
        lon /= 1_000_000
    return pd.Series([lat, lon])

def clean_geojson(gdf, output_path):  # Save GeoDataFrame in a clean format GeoJSON file.
    temp_path = output_path.replace(".geojson", "_temp.geojson")
    gdf.to_file(temp_path, driver="GeoJSON")

    with open(temp_path, "r", encoding="utf-8") as file:
        geojson_data = json.load(file)

    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(geojson_data, file, indent=4)

    print(f"✅ Saved: {output_path}")

def find_coordinate_columns(df, prefix, column_type): 
    """
    Find coordinate columns with various naming patterns
    
    df: DataFrame to search
    prefix: 'start' or 'end' to indicate which endpoint to find
    column_type: 'lat' or 'lon' to indicate latitude or longitude
    
    Returns: Column name if found, None otherwise
    """
    patterns = []
    
    if prefix == 'start':
        if column_type == 'lat':
            patterns = [
                'awal latitude', 'awal lat', 
                'koordinat latitude', 'koordinat lat',
                'latitude awal', 'lat awal',
                'latitude 1', 'lat 1'
            ]
        else:  # longitude
            patterns = [
                'awal longitude', 'awal lon', 
                'koordinat longitude', 'koordinat lon',
                'longitude awal', 'lon awal',
                'longitude 1', 'lon 1'
            ]
    else:  # end
        if column_type == 'lat':
            patterns = [
                'akhir latitude', 'akhir lat',
                'latitude akhir', 'lat akhir',
                'latitude 2', 'lat 2'
            ]
        else:  # longitude
            patterns = [
                'akhir longitude', 'akhir lon',
                'longitude akhir', 'lon akhir',
                'longitude 2', 'lon 2'
            ]
    
    # If we're looking for start coordinates, and just a single coordinate exists (no start/end distinction)
    if prefix == 'start' and column_type == 'lat':
        # Add cases where latitude exists without start/end specifier
        patterns.extend(['latitude', 'lat'])
    elif prefix == 'start' and column_type == 'lon':
        # Add cases where longitude exists without start/end specifier
        patterns.extend(['longitude', 'lon'])
    
    # Search for column containing any of the patterns
    for pattern in patterns:
        for col in df.columns:
            if isinstance(col, str) and pattern in col.lower():
                return col
    
    return None

def flatten_excel_to_geojson(file_path, output_folder): #Convert all sheets from an Excel file to GeoJSON, handling both Point and LineString geometries

    # Load workbook
    wb = load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge cells and fill values
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(str(merge))
            top_left = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    ws.cell(row, col, top_left)

        # Convert to DataFrame
        data = list(ws.values)
        df = pd.DataFrame(data)

        # Check if DataFrame is empty or has too few rows
        if df.empty or len(df) < 3:
            print(f"⚠️ Skipping '{sheet_name}' (Empty sheet or insufficient data)")
            continue

        # Find rows containing "NO" (safer approach)
        try:
            header_indices = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index
            if len(header_indices) == 0:
                print(f"⚠️ Skipping '{sheet_name}' (No header row with 'NO' found)")
                continue
            header_index = header_indices[0]
        except Exception as e:
            print(f"⚠️ Skipping '{sheet_name}' (Error finding header row: {str(e)})")
            continue

        # Use the identified header row
        df.columns = df.iloc[header_index].astype(str).str.strip()

        # Remove empty columns
        df = df.dropna(axis=1, how="all")

        # First pass of REKAP filtering - do this early and more aggressively
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} REKAP columns in first pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns - {str(e)}")

        # Drop rows above and including header, plus the empty row after header
        rows_to_drop = list(range(0, header_index + 2))
        if len(df) > max(rows_to_drop) + 1:  # Make sure we have enough rows
            df = df.drop(index=rows_to_drop).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough data rows after header)")
            continue

        # Improved header row merging
        if len(df) >= 2:
            # Convert all values to strings to handle None values safely
            first_row = df.iloc[0].astype(str).replace('None', '').replace('nan', '')
            second_row = df.iloc[1].astype(str).replace('None', '').replace('nan', '')
            
            # Smart merging to avoid duplication
            merged_header = []
            for a, b in zip(first_row, second_row):
                a = a.strip()
                b = b.strip()
                
                # Skip columns with "REKAP" in the name - additional check
                if "rekap" in a.lower() or "rekap" in b.lower():
                    merged_header.append("TO_BE_REMOVED")  # Mark for removal
                    continue
                
                if not a and not b:
                    merged_header.append("Column_" + str(len(merged_header)))
                elif not a:
                    merged_header.append(b)
                elif not b:
                    merged_header.append(a)
                else:
                    # Check if one is contained in the other to avoid redundancy
                    if a.lower() in b.lower():
                        merged_header.append(b)
                    elif b.lower() in a.lower():
                        merged_header.append(a)
                    else:
                        merged_header.append(f"{a} {b}")
            
            # Ensure column names are unique
            df.columns = unique_column_names(merged_header)
            
            # Remove any columns marked for removal
            df = df.loc[:, ~df.columns.str.contains("TO_BE_REMOVED")]
            
            # Remove the first two rows used for headers
            df = df.drop(index=[0, 1]).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough rows for headers)")
            continue

        # Second pass of REKAP filtering after merging headers
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} REKAP columns in second pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns (second pass) - {str(e)}")

        # Normalize column names for consistent detection (safely handle None values)
        df.columns = [str(col).lower().strip() if col is not None else f"col_{i}" 
                      for i, col in enumerate(df.columns)]
        
        # Third pass of REKAP filtering after normalizing
        try:
            rekap_mask = df.columns.str.contains("rekap", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} rekap columns in third pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering rekap columns (third pass) - {str(e)}")
        
        # Safely check if this sheet contains "PAGAR PENGAMAN" or "MARKA" columns
        contains_linestring_indicator = any(col for col in df.columns 
                                 if isinstance(col, str) and 
                                 ("pagar pengaman" in col.lower() or "marka" in col.lower()))

        # Find all coordinate columns using the improved function
        start_lat_col = find_coordinate_columns(df, 'start', 'lat')
        start_lon_col = find_coordinate_columns(df, 'start', 'lon')
        end_lat_col = find_coordinate_columns(df, 'end', 'lat')
        end_lon_col = find_coordinate_columns(df, 'end', 'lon')
        
        # Check available coordinate patterns
        has_start_coords = start_lat_col is not None and start_lon_col is not None
        has_end_coords = end_lat_col is not None and end_lon_col is not None

        if not has_start_coords:
            print(f"⚠️ Skipping '{sheet_name}' (No valid start coordinate columns found)")
            continue

        # Convert all coordinate columns to numeric
        if has_start_coords:
            df[start_lat_col] = pd.to_numeric(df[start_lat_col], errors='coerce')
            df[start_lon_col] = pd.to_numeric(df[start_lon_col], errors='coerce')
            
            # Fix coordinates if needed
            df[[start_lat_col, start_lon_col]] = df.apply(fix_coordinates, axis=1, lat_col=start_lat_col, lon_col=start_lon_col)
        
        if has_end_coords:
            df[end_lat_col] = pd.to_numeric(df[end_lat_col], errors='coerce')
            df[end_lon_col] = pd.to_numeric(df[end_lon_col], errors='coerce')
            
            # Fix coordinates if needed
            df[[end_lat_col, end_lon_col]] = df.apply(fix_coordinates, axis=1, lat_col=end_lat_col, lon_col=end_lon_col)
        
        # Check if the end coordinates actually contain valid data
        if has_end_coords:
            has_valid_end_coords = not df[end_lat_col].isna().all() and not df[end_lon_col].isna().all()
            
            # Check if we have at least one row with both coordinates valid
            valid_pairs = ((df[start_lat_col].notna() & df[start_lon_col].notna()) & 
                           (df[end_lat_col].notna() & df[end_lon_col].notna())).any()
            
            # Only use LineString if we have valid end coordinates and at least one pair is complete
            use_linestring = has_valid_end_coords and valid_pairs and contains_linestring_indicator
        else:
            use_linestring = False
        
        # Determine geometry type based on available coordinates and actual data
        if use_linestring:
            # LineString geometry with start/end points
            print(f"Processing '{sheet_name}' as LineString (both start and end coordinates contain valid data)")
            
            # Create LineString geometry
            df["geometry"] = df.apply(
                lambda row: LineString([
                    (row[start_lon_col], row[start_lat_col]),
                    (row[end_lon_col], row[end_lat_col])
                ]) if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) and 
                     pd.notna(row[end_lon_col]) and pd.notna(row[end_lat_col]) else None,
                axis=1
            )
            
            # Exclude coordinate columns from properties
            exclude_cols = [start_lat_col, start_lon_col, end_lat_col, end_lon_col, "geometry"]
            
        else:
            # Point geometry (only start coordinates)
            print(f"Processing '{sheet_name}' as Point geometry (second coordinates are missing or invalid)")
            
            # Create Point geometry with start coordinates
            df["geometry"] = df.apply(
                lambda row: Point(row[start_lon_col], row[start_lat_col]) 
                if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) else None,
                axis=1
            )
            
            # Exclude coordinate columns from properties
            exclude_cols = [start_lat_col, start_lon_col]
            if has_end_coords:
                exclude_cols.extend([end_lat_col, end_lon_col])
            exclude_cols.append("geometry")
        
        # Drop rows where geometry is None
        df = df.dropna(subset=["geometry"]).reset_index(drop=True)

        # Only create GeoDataFrame if there are valid geometries
        if len(df) > 0 and not df["geometry"].isnull().all():
            # Fourth pass of REKAP filtering
            # Filter REKAP columns before creating properties
            properties_cols = [col for col in df.columns if col not in exclude_cols]
            properties_cols = [col for col in properties_cols if 'rekap' not in str(col).lower()]
            
            gdf = gpd.GeoDataFrame(df[properties_cols + ["geometry"]], crs="EPSG:4326")
            
            # Apply column renaming after creating the GeoDataFrame
            gdf.columns = clean_column_names(gdf.columns)

            # Fifth pass REKAP filter - case insensitive filter after cleaning column names
            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
            
            # Remove unwanted "None_" and "None" columns
            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.match(r"^None$|None_", na=False)]

            # Remove " None" from remaining column names
            gdf.columns = gdf.columns.astype(str).str.replace(r"\sNone\b", "", regex=True).str.strip()

            # Create output folder if it doesn't exist
            os.makedirs(output_folder, exist_ok=True)

            # Define output file path
            output_path = os.path.join(output_folder, f"{sheet_name}.geojson")

            # One final check for REKAP columns before saving
            if any('rekap' in str(col).lower() for col in gdf.columns):
                print(f"Warning: Some REKAP columns still exist after all filtering: {[col for col in gdf.columns if 'rekap' in str(col).lower()]}")
                # Last resort: manually drop these columns
                gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]

            # Save the GeoJSON in a clean format
            clean_geojson(gdf, output_path)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (No valid geometry found)")
            continue  # Skip processing this sheet if no valid geometries exist

def process_excel_folder(input_folder, output_base_folder): #Process all Excel files in a folder and convert them to GeoJSON
    
    # Create the 'Extract GeoJSON' folder inside the output folder
    output_folder = os.path.join(output_base_folder, "Extract GeoJSON")
    os.makedirs(output_folder, exist_ok=True)
    
    # Get all Excel files in the input folder
    excel_extensions = ['*.xlsx', '*.xls', '*.xlsm']
    excel_files = []
    
    for ext in excel_extensions:
        excel_files.extend(glob.glob(os.path.join(input_folder, ext)))
    
    if not excel_files:
        print(f"⚠️ No Excel files found in {input_folder}")
        return
    
    # Process each Excel file
    for i, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print(f"\n[{i}/{len(excel_files)}] Processing: {file_name}")
        
        # Create a subfolder for each Excel file
        file_output_folder = os.path.join(output_folder, Path(file_name).stem)
        os.makedirs(file_output_folder, exist_ok=True)
        
        try:
            flatten_excel_to_geojson(file_path, file_output_folder)
            print(f"✓ Completed processing: {file_name}")
        except Exception as e:
            print(f"❌ Error processing {file_name}: {str(e)}")
    
    print(f"\n🎉 All Excel files processed. Output saved to: {output_folder}")
    
    # Clean up temporary files
    cleanup_temp_files(output_folder)

# Delete temporary files
def cleanup_temp_files(output_folder):
    for temp_file in glob.glob(os.path.join(output_folder, "**/*_temp.geojson"), recursive=True):
        try:
            os.remove(temp_file)
            print(f"Removed temporary file: {temp_file}")
        except Exception as e:
            print(f"Error removing temporary file {temp_file}: {str(e)}")

# %% [markdown]
# ### 2.2. Run Function Excel to GeoJSON

# %%
input_folder = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\Excel Folder"  # Fill with the path file of excel
output_base_folder = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\Check GeoJSON"  # Fill with the path folder of export result
process_excel_folder(input_folder, output_base_folder) # Run the function!

# %% [markdown]
# ## 3. Application to Export Excel into Shapefile

# %% [markdown]
# ### 3.1. Function Codes

# %%
def unique_column_names(columns): # Ensure column names are unique by appending suffix.
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def clean_column_names(columns): # Standardize column names by capitalizing each word properly.
    cleaned_columns = []
    seen = {}
    for col in columns:
        col = str(col).strip()
        
        # Remove columns with "Rekap" in the name
        if "rekap" in col.lower():
            continue
        
        # Fix for duplicated words (like "No No", "Detail Lokasi Detail Lokasi")
        words = col.split()
        if len(words) >= 2:
            # Check for repeated word patterns
            half_length = len(words) // 2
            if words[:half_length] == words[half_length:] and len(words) % 2 == 0:
                # If the first half matches the second half, only use the first half
                col = " ".join(words[:half_length])
        
        # Capitalize words
        col = " ".join(word.capitalize() for word in col.split())
        
        # Handle duplicates
        if col in seen:
            seen[col] += 1
            col = f"{col} {seen[col]}"
        else:
            seen[col] = 0
        
        # Shapefile has column name length limitation (10 chars)
        # We'll handle this by truncating to 10 chars while ensuring uniqueness
        col = col[:10]
        
        cleaned_columns.append(col)
    return cleaned_columns

def fix_coordinates(row, lat_col, lon_col): #Fix latitude and longitude values that may be in the wrong format.
    lat, lon = row[lat_col], row[lon_col]
    if pd.notna(lat) and abs(lat) > 90:
        lat /= 1_000_000
    if pd.notna(lon) and abs(lon) > 180:
        lon /= 1_000_000
    return pd.Series([lat, lon])

def save_shapefile(gdf, output_path): 
    """
    Save GeoDataFrame to Shapefile format with warnings suppressed
    
    gdf: GeoDataFrame to save
    output_path: Path where to save the shapefile
    """
    try:
        # Truncate column names to 10 characters ourselves to avoid warnings
        # This proactively handles the limitation rather than relying on geopandas/pyogrio to do it
        orig_columns = gdf.columns.tolist()
        truncated_columns = [str(col)[:10] for col in orig_columns]
        
        # Create a mapping dictionary to show what changed
        column_mapping = {orig: trunc for orig, trunc in zip(orig_columns, truncated_columns)}
        
        # Check for duplicates after truncation and make them unique
        seen = {}
        unique_truncated = []
        for col in truncated_columns:
            if col in seen:
                seen[col] += 1
                unique_truncated.append(f"{col[:7]}_{seen[col]}")
            else:
                seen[col] = 0
                unique_truncated.append(col)
        
        # Rename the columns with the unique truncated names
        gdf.columns = unique_truncated
        
        # Import warnings to suppress them
        import warnings
        
        # Suppress specific warnings
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, message='.*truncated when saved to ESRI Shapefile.*')
            warnings.filterwarnings('ignore', category=RuntimeWarning, message='.*Normalized/laundered field name.*')
            
            # Save to shapefile with warnings suppressed
            gdf.to_file(output_path, driver="ESRI Shapefile")
        
        print(f"✅ Saved: {output_path}")
                
    except Exception as e:
        print(f"❌ Error saving shapefile {output_path}: {str(e)}")

def find_coordinate_columns(df, prefix, column_type):
    """
    Find coordinate columns with various naming patterns
    
    df: DataFrame to search
    prefix: 'start' or 'end' to indicate which endpoint to find
    column_type: 'lat' or 'lon' to indicate latitude or longitude
    
    Returns Column name if found, None otherwise
    """
    patterns = []
    
    if prefix == 'start':
        if column_type == 'lat':
            patterns = [
                'awal latitude', 'awal lat', 
                'koordinat latitude', 'koordinat lat',
                'latitude awal', 'lat awal',
                'latitude 1', 'lat 1'
            ]
        else:  # longitude
            patterns = [
                'awal longitude', 'awal lon', 
                'koordinat longitude', 'koordinat lon',
                'longitude awal', 'lon awal',
                'longitude 1', 'lon 1'
            ]
    else:  # end
        if column_type == 'lat':
            patterns = [
                'akhir latitude', 'akhir lat',
                'latitude akhir', 'lat akhir',
                'latitude 2', 'lat 2'
            ]
        else:  # longitude
            patterns = [
                'akhir longitude', 'akhir lon',
                'longitude akhir', 'lon akhir',
                'longitude 2', 'lon 2'
            ]
    
    # If we're looking for start coordinates, and just a single coordinate exists (no start/end distinction)
    if prefix == 'start' and column_type == 'lat':
        # Add cases where latitude exists without start/end specifier
        patterns.extend(['latitude', 'lat'])
    elif prefix == 'start' and column_type == 'lon':
        # Add cases where longitude exists without start/end specifier
        patterns.extend(['longitude', 'lon'])
    
    # Search for column containing any of the patterns
    for pattern in patterns:
        for col in df.columns:
            if isinstance(col, str) and pattern in col.lower():
                return col
    
    return None

def flatten_excel_to_shapefile(file_path, output_folder): #Convert all sheets from an Excel file to Shapefile, handling both Point and LineString geometries

    # Load workbook
    wb = load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge cells and fill values
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(str(merge))
            top_left = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    ws.cell(row, col, top_left)

        # Convert to DataFrame
        data = list(ws.values)
        df = pd.DataFrame(data)

        # Check if DataFrame is empty or has too few rows
        if df.empty or len(df) < 3:
            print(f"⚠️ Skipping '{sheet_name}' (Empty sheet or insufficient data)")
            continue

        # Find rows containing "NO" (safer approach)
        try:
            header_indices = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index
            if len(header_indices) == 0:
                print(f"⚠️ Skipping '{sheet_name}' (No header row with 'NO' found)")
                continue
            header_index = header_indices[0]
        except Exception as e:
            print(f"⚠️ Skipping '{sheet_name}' (Error finding header row: {str(e)})")
            continue

        # Use the identified header row
        df.columns = df.iloc[header_index].astype(str).str.strip()

        # Remove empty columns
        df = df.dropna(axis=1, how="all")

        # First pass of REKAP filtering - do this early and more aggressively
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} REKAP columns in first pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns - {str(e)}")

        # Drop rows above and including header, plus the empty row after header
        rows_to_drop = list(range(0, header_index + 2))
        if len(df) > max(rows_to_drop) + 1:  # Make sure we have enough rows
            df = df.drop(index=rows_to_drop).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough data rows after header)")
            continue

        # Improved header row merging
        if len(df) >= 2:
            # Convert all values to strings to handle None values safely
            first_row = df.iloc[0].astype(str).replace('None', '').replace('nan', '')
            second_row = df.iloc[1].astype(str).replace('None', '').replace('nan', '')
            
            # Smart merging to avoid duplication
            merged_header = []
            for a, b in zip(first_row, second_row):
                a = a.strip()
                b = b.strip()
                
                # Skip columns with "REKAP" in the name - additional check
                if "rekap" in a.lower() or "rekap" in b.lower():
                    merged_header.append("TO_BE_REMOVED")  # Mark for removal
                    continue
                
                if not a and not b:
                    merged_header.append("Column_" + str(len(merged_header)))
                elif not a:
                    merged_header.append(b)
                elif not b:
                    merged_header.append(a)
                else:
                    # Check if one is contained in the other to avoid redundancy
                    if a.lower() in b.lower():
                        merged_header.append(b)
                    elif b.lower() in a.lower():
                        merged_header.append(a)
                    else:
                        merged_header.append(f"{a} {b}")
            
            # Ensure column names are unique
            df.columns = unique_column_names(merged_header)
            
            # Remove any columns marked for removal
            df = df.loc[:, ~df.columns.str.contains("TO_BE_REMOVED")]
            
            # Remove the first two rows used for headers
            df = df.drop(index=[0, 1]).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough rows for headers)")
            continue

        # Second pass of REKAP filtering after merging headers
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} REKAP columns in second pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns (second pass) - {str(e)}")

        # Normalize column names for consistent detection (safely handle None values)
        df.columns = [str(col).lower().strip() if col is not None else f"col_{i}" 
                      for i, col in enumerate(df.columns)]
        
        # Third pass of REKAP filtering after normalizing
        try:
            rekap_mask = df.columns.str.contains("rekap", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
                print(f"Removed {rekap_mask.sum()} rekap columns in third pass")
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering rekap columns (third pass) - {str(e)}")
        
        # Safely check if this sheet contains "PAGAR PENGAMAN" or "MARKA" columns
        contains_linestring_indicator = any(col for col in df.columns 
                                 if isinstance(col, str) and 
                                 ("pagar pengaman" in col.lower() or "marka" in col.lower()))

        # Find all coordinate columns using the improved function
        start_lat_col = find_coordinate_columns(df, 'start', 'lat')
        start_lon_col = find_coordinate_columns(df, 'start', 'lon')
        end_lat_col = find_coordinate_columns(df, 'end', 'lat')
        end_lon_col = find_coordinate_columns(df, 'end', 'lon')
        
        # Check available coordinate patterns
        has_start_coords = start_lat_col is not None and start_lon_col is not None
        has_end_coords = end_lat_col is not None and end_lon_col is not None

        if not has_start_coords:
            print(f"⚠️ Skipping '{sheet_name}' (No valid start coordinate columns found)")
            continue

        # Convert all coordinate columns to numeric
        if has_start_coords:
            df[start_lat_col] = pd.to_numeric(df[start_lat_col], errors='coerce')
            df[start_lon_col] = pd.to_numeric(df[start_lon_col], errors='coerce')
            
            # Fix coordinates if needed
            df[[start_lat_col, start_lon_col]] = df.apply(fix_coordinates, axis=1, lat_col=start_lat_col, lon_col=start_lon_col)
        
        if has_end_coords:
            df[end_lat_col] = pd.to_numeric(df[end_lat_col], errors='coerce')
            df[end_lon_col] = pd.to_numeric(df[end_lon_col], errors='coerce')
            
            # Fix coordinates if needed
            df[[end_lat_col, end_lon_col]] = df.apply(fix_coordinates, axis=1, lat_col=end_lat_col, lon_col=end_lon_col)
        
        # Check if the end coordinates actually contain valid data
        if has_end_coords:
            has_valid_end_coords = not df[end_lat_col].isna().all() and not df[end_lon_col].isna().all()
            
            # Check if we have at least one row with both coordinates valid
            valid_pairs = ((df[start_lat_col].notna() & df[start_lon_col].notna()) & 
                           (df[end_lat_col].notna() & df[end_lon_col].notna())).any()
            
            # Only use LineString if we have valid end coordinates and at least one pair is complete
            use_linestring = has_valid_end_coords and valid_pairs and contains_linestring_indicator
        else:
            use_linestring = False
        
        # Determine geometry type based on available coordinates and actual data
        if use_linestring:
            # LineString geometry with start/end points
            print(f"Processing '{sheet_name}' as LineString (both start and end coordinates contain valid data)")
            
            # Create LineString geometry
            df["geometry"] = df.apply(
                lambda row: LineString([
                    (row[start_lon_col], row[start_lat_col]),
                    (row[end_lon_col], row[end_lat_col])
                ]) if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) and 
                     pd.notna(row[end_lon_col]) and pd.notna(row[end_lat_col]) else None,
                axis=1
            )
            
            # Exclude coordinate columns from properties
            exclude_cols = [start_lat_col, start_lon_col, end_lat_col, end_lon_col, "geometry"]
            
        else:
            # Point geometry (only start coordinates)
            print(f"Processing '{sheet_name}' as Point geometry (second coordinates are missing or invalid)")
            
            # Create Point geometry with start coordinates
            df["geometry"] = df.apply(
                lambda row: Point(row[start_lon_col], row[start_lat_col]) 
                if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) else None,
                axis=1
            )
            
            # Exclude coordinate columns from properties
            exclude_cols = [start_lat_col, start_lon_col]
            if has_end_coords:
                exclude_cols.extend([end_lat_col, end_lon_col])
            exclude_cols.append("geometry")
        
        # Drop rows where geometry is None
        df = df.dropna(subset=["geometry"]).reset_index(drop=True)

        # Only create GeoDataFrame if there are valid geometries
        if len(df) > 0 and not df["geometry"].isnull().all():
            # Fourth pass of REKAP filtering
            # Filter REKAP columns before creating properties
            properties_cols = [col for col in df.columns if col not in exclude_cols]
            properties_cols = [col for col in properties_cols if 'rekap' not in str(col).lower()]
            
            gdf = gpd.GeoDataFrame(df[properties_cols + ["geometry"]], crs="EPSG:4326")
            
            # Apply column renaming after creating the GeoDataFrame
            gdf.columns = clean_column_names(gdf.columns)

            # Fifth pass REKAP filter - case insensitive filter after cleaning column names
            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
            
            # Remove unwanted "None_" and "None" columns
            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.match(r"^None$|None_", na=False)]

            # Remove " None" from remaining column names
            gdf.columns = gdf.columns.astype(str).str.replace(r"\sNone\b", "", regex=True).str.strip()

            # Create output folder if it doesn't exist
            os.makedirs(output_folder, exist_ok=True)

            # Define output file path - changed from .geojson to .shp
            output_path = os.path.join(output_folder, f"{sheet_name}.shp")

            # One final check for REKAP columns before saving
            if any('rekap' in str(col).lower() for col in gdf.columns):
                print(f"Warning: Some REKAP columns still exist after all filtering: {[col for col in gdf.columns if 'rekap' in str(col).lower()]}")
                # Last resort: manually drop these columns
                gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]

            # Save the Shapefile using our new save_shapefile function
            save_shapefile(gdf, output_path)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (No valid geometry found)")
            continue  # Skip processing this sheet if no valid geometries exist

def process_excel_folder(input_folder, output_base_folder): # Process all Excel files in a folder and convert them to Shapefiles
    
    # Create the 'Extract Shapefile' folder inside the output folder - changed from 'Extract GeoJSON'
    output_folder = os.path.join(output_base_folder, "Extract Shapefile")
    os.makedirs(output_folder, exist_ok=True)
    
    # Get all Excel files in the input folder
    excel_extensions = ['*.xlsx', '*.xls', '*.xlsm']
    excel_files = []
    
    for ext in excel_extensions:
        excel_files.extend(glob.glob(os.path.join(input_folder, ext)))
    
    if not excel_files:
        print(f"⚠️ No Excel files found in {input_folder}")
        return
    
    # Process each Excel file
    for i, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print(f"\n[{i}/{len(excel_files)}] Processing: {file_name}")
        
        # Create a subfolder for each Excel file
        file_output_folder = os.path.join(output_folder, Path(file_name).stem)
        os.makedirs(file_output_folder, exist_ok=True)
        
        try:
            flatten_excel_to_shapefile(file_path, file_output_folder)
            print(f"✓ Completed processing: {file_name}")
        except Exception as e:
            print(f"❌ Error processing {file_name}: {str(e)}")
    
    print(f"\n🎉 All Excel files processed. Output saved to: {output_folder}")

# %% [markdown]
# ### 3.2. Run Function Excel to Shapefile

# %%
input_folder = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\Excel Folder"  # Fill with the path file of excel
output_base_folder = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\Check GeoJSON"  # Fill with the path folder of export result
process_excel_folder(input_folder, output_base_folder) # Run the function!


