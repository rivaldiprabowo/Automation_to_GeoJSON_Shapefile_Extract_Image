# %% [markdown]
# # Application to Extract Image from Excel

# %% [markdown]
# ## 1. Import Library

# %%
import os
import pandas as pd
import numpy as np
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image
import io

# %%
def unique_column_names(columns): #Ensure column names are unique by appending a suffix.
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def extract_images_from_excel(file_path, output_folder): #Extract images from columns containing 'DOKUMENTASI' in all sheets and save them.
    
    wb = load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        image_loader = SheetImageLoader(ws)  # Load images from the sheet
        
        # Unmerge cells and fill values
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(str(merge))
            top_left = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    ws.cell(row, col, top_left)
        
        # Convert to DataFrame
        data = list(ws.values)
        df = pd.DataFrame(data)
        
        # Identify the header row
        try:
            header_index = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index[0]
        except IndexError:
            print(f"⚠️ Could not identify header row in sheet '{sheet_name}', skipping...")
            continue
        
        # Set the header
        df.columns = df.iloc[2].astype(str).str.strip()
        
        # Remove empty columns
        df = df.dropna(axis=1, how="all")
        
        # Drop "REKAP" section if present
        df = df.loc[:, ~df.columns.str.contains("REKAP", case=False, na=False)]
        df = df.drop(index=[0, 1, 4]).reset_index(drop=True)
        
        # Merge first two rows if needed
        merged_header = [a if a == b else f"{a} {b}" for a, b in zip(df.iloc[0], df.iloc[1])]
        
        # Ensure column names are unique
        df.columns = unique_column_names(merged_header)
        
        # Remove the first two rows used for headers
        df = df.drop(index=[0, 1]).reset_index(drop=True)
        
        # Normalize column names for consistent detection
        df.columns = df.columns.str.upper().str.strip()
        
        # Find all columns that contain "DOKUMENTASI" (case-insensitive)
        dokumentasi_cols = [col for col in df.columns if "DOKUMENTASI" in col.upper()]
        
        # Create output directory
        sheet_folder = os.path.join(output_folder, sheet_name)
        os.makedirs(sheet_folder, exist_ok=True)
        
        # Extract images
        for row_idx in range(2, ws.max_row + 1):  # Assuming images are below the header row
            for col_idx, col_name in enumerate(df.columns):
                if col_name not in dokumentasi_cols:
                    continue
                
                cell_address = ws.cell(row=row_idx, column=col_idx + 1).coordinate  # Get Excel cell address (e.g., B5)
                
                if image_loader.image_in(cell_address):  # Check if image exists in the cell
                    try:
                        img = image_loader.get(cell_address)  # Get the image object
                        
                        # Create a new BytesIO object and keep it open until we're done using it
                        img_bytes = io.BytesIO()
                        img.save(img_bytes, format="PNG")
                        img_bytes.seek(0)  # Reset pointer
                        
                        # Save the image
                        img_path = os.path.join(sheet_folder, f"{sheet_name}_Column {col_name}_Row {row_idx}.png")
                        with open(img_path, 'wb') as f:
                            f.write(img_bytes.getvalue())
                        
                        print(f"✅ Image saved: {img_path}")
                    except Exception as e:
                        print(f"❌ Error saving image at {cell_address}: {str(e)}")
    
    print("🎉 Image extraction completed!")

# %%
excel_file = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\01. Cileungsi - Cibeet.xlsx"  # Fill with the path file of excel
export_folder = r"C:\Users\kanzi\Documents\Part Time Job\Automation Codes\check_photo"  # Fill with the path folder of export result
extract_images_from_excel(excel_file, export_folder) # Run the function!


