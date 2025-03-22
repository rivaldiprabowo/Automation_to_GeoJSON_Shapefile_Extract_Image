import sys
import subprocess
import pkg_resources
import json
import glob
import os
import re
import io
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import geopandas as gpd
from datetime import datetime
from shapely.geometry import Point, LineString, MultiPoint
import geopandas as gpd
import numpy as np
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter


class ExcelConverter:
    def __init__(self, output_folder, log_callback = None, progress_callback = None) -> None:
        self.output_folder = output_folder
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.list_df = {}

    def _update_progress(self, percent):
        if self.progress_callback:
            self.progress_callback(percent)

    def _log(self, message):
        if self.log_callback:
            self.log_callback(message)

    def install_requirements(): #Install required packages if they are not already installed.
        required = {
            'pandas': '1.0.0',
            'openpyxl': '3.0.0',
            'geopandas': '0.9.0',
            'shapely': '1.7.0',}
        
        installed = {pkg.key: pkg.version for pkg in pkg_resources.working_set}
        
        # Determine what needs to be installed
        missing = []
        update = []
        
        for package, min_version in required.items():
            if package not in installed:
                missing.append(package)
            elif pkg_resources.parse_version(installed[package]) < pkg_resources.parse_version(min_version):
                update.append(package)
        
        if missing or update: # If packages need to be installed or updated
            print("Some required packages are missing or need to be updated.")
            print(f"Missing: {', '.join(missing) if missing else 'None'}")
            print(f"Need update: {', '.join(update) if update else 'None'}")
            
            try:  # Install missing packages
                if missing:
                    print(f"Installing missing packages: {', '.join(missing)}")
                    subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)

                if update: # Update packages that need updating
                    print(f"Updating packages: {', '.join(update)}")
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade"] + update)
                    
                print("All required packages have been installed/updated successfully!")
                
                # Re-import the modules to make them available
                if 'pandas' in missing or 'pandas' in update:
                    global pd
                    import pandas as pd
                if 'openpyxl' in missing or 'openpyxl' in update:
                    global load_workbook
                    from openpyxl import load_workbook
                if 'geopandas' in missing or 'geopandas' in update:
                    global gpd
                    import geopandas as gpd
                if 'shapely' in missing or 'shapely.geometry' in update:
                    global Point, LineString
                    from shapely.geometry import Point, LineString
                    
            except Exception as e:
                print(f"Failed to install required packages: {str(e)}")
                sys.exit(1)

    def unique_column_names(columns): #Ensure column names are unique by appending suffix.
        seen = {}
        new_columns = []
        for col in columns:
            if col is None:
                col = "Unnamed"
            
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_columns.append(col)
        return new_columns

    def clean_column_names(columns): #Standardize column names by capitalizing each word properly.
        cleaned_columns = []
        seen = {}
        for col in columns:
            if col is None:
                col = "Unnamed"
            else:
                col = str(col).strip()
                
            if "rekap" in col.lower(): # Remove columns with "Rekap" in the name
                continue

            words = col.split() # Fix for duplicated words (like "No No", "Detail Lokasi Detail Lokasi")
            if len(words) >= 2:
                half_length = len(words) // 2 # Check for repeated word patterns
                if words[:half_length] == words[half_length:] and len(words) % 2 == 0:
                    col = " ".join(words[:half_length]) # If the first half matches the second half, only use the first half
            
            col = " ".join(word.capitalize() for word in col.split())
            
            # Handle duplicates
            if col in seen:
                seen[col] += 1
                col = f"{col} {seen[col]}"
            else:
                seen[col] = 0
            
            cleaned_columns.append(col)
        return cleaned_columns

    def log_coordinate_errors(error_logs, output_base_folder): # Logs to store error in coordinate data
        if not error_logs:
            print("No coordinate errors to log.")
            return
            
        # Create a DataFrame from the error logs with better formatting
        df = pd.DataFrame(error_logs)
        
        # Create a timestamp for the log file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Define the log file path
        log_path = os.path.join(output_base_folder, f"Coordinate_Error_Log_{timestamp}.xlsx")
        
        # Add a summary column for easier reviewing
        if not df.empty:
            # Create a summary column that combines the key information
            df['Error Summary'] = df.apply(
                lambda row: f"File: {row['Excel File']}, Sheet: {row['Sheet']}, " + 
                        f"Row: {row['Row Index']}, Error: {row['Error']}", 
                axis=1
            )
        
        # Write the DataFrame to an Excel file
        try:
            with pd.ExcelWriter(log_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Coordinate Errors', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Coordinate Errors']
                for i, col in enumerate(df.columns):
                    max_length = max(df[col].astype(str).map(len).max(), len(col))
                    worksheet.column_dimensions[chr(65 + i)].width = max_length + 2
                    
            print(f"✅ Error log saved to: {log_path}")
        except Exception as e:
            print(f"❌ Error saving coordinate error log: {str(e)}")

    def find_coordinate_columns(df, prefix, column_type): #Find coordinate columns with various naming patterns returns column name
        
        patterns = []
        
        if prefix == 'start':
            if column_type == 'lat':
                patterns = ['awal latitude','awal lat','koordinat latitude','koordinat lat','latitude awal','lat awal','latitude 1','lat 1']
            else:
                patterns = ['awal longitude','awal lon','koordinat longitude','koordinat lon','longitude awal', 'lon awal','longitude 1','lon 1' ]
        else:
            if column_type == 'lat':
                patterns = ['akhir latitude','akhir lat','latitude akhir','lat akhir','latitude 2','lat 2']
            else:
                patterns = ['akhir longitude','akhir lon','longitude akhir','lon akhir','longitude 2','lon 2']
        
        # If we're looking for start coordinates, and just a single coordinate exists (no start/end distinction)
        if prefix == 'start' and column_type == 'lat':
            patterns.extend(['latitude', 'lat'])
        elif prefix == 'start' and column_type == 'lon':
            patterns.extend(['longitude', 'lon'])
        
        # Search for column containing any of the patterns
        for pattern in patterns:
            for col in df.columns:
                if col is not None and isinstance(col, str) and pattern in col.lower():
                    return col
        
        return None

    def fix_coordinates(row, lat_col, lon_col): #Fix latitude and longitude values that may be in the wrong format.
        # Defensive check to make sure columns exist in the row
        if lat_col not in row or lon_col not in row:
            return pd.Series([pd.NA, pd.NA])
            
        lat, lon = row[lat_col], row[lon_col]
        original_lat, original_lon = lat, lon 
        
        # Handle string values with commas (e.g., "-69,694,951" or "1,065,663,308")
        if isinstance(lat, str):
            try:
                lat = float(lat.replace(',', ''))
            except (ValueError, AttributeError):
                lat = pd.NA
        
        if isinstance(lon, str):
            try:
                lon = float(lon.replace(',', ''))
            except (ValueError, AttributeError):
                lon = pd.NA
        
        # First attempt to detect the scale by number of digits
        if pd.notna(lat):
            lat_abs = abs(lat)
            # For values like -6448977 (which should be around -6.4 degrees)
            if 1_000_000 < lat_abs < 10_000_000 and str(int(lat_abs)).startswith(('6', '7', '8', '9')):
                lat = lat / 1_000_000
            elif lat_abs > 90:
                if lat_abs > 10_000_000:  # Very large values
                    lat = lat / 10_000_000
                elif lat_abs > 1_000_000:  # Large values (common for Indonesia coords)
                    lat = lat / 1_000_000
                elif lat_abs > 90_000:
                    lat = lat / 1_000
        
        if pd.notna(lon):
            lon_abs = abs(lon)
            if 100_000_000 < lon_abs < 1_500_000_000:
                lon = lon / 10_000_000
            elif lon_abs > 180:
                if lon_abs > 10_000_000:
                    lon = lon / 10_000_000
                elif lon_abs > 1_000_000:
                    lon = lon / 1_000_000
                elif lon_abs > 180_000:
                    lon = lon / 1_000
        
        return pd.Series([lat, lon])

    def parse_coordinate(coord_value): # Parse and clean coordinate values that may be stored in unwanted formats

        if pd.isna(coord_value):
            return pd.NA
        
        # If already numeric, return as is
        if pd.api.types.is_numeric_dtype(type(coord_value)):
            return float(coord_value)
        
        # Handle coordinate in string values
        if isinstance(coord_value, str):
            coord_string = coord_value.strip()

            # Case 1: Check if it's a simple decimal with apostrophes and try simple conversion
            cleaned = coord_string.replace("'", "")
            cleaned = cleaned.replace(",", ".")
            
            try:
                return float(cleaned)
            except ValueError:
                pass
            
            # Case 2: Parse DMS format like "107°18'40.74"E" or "6°17'23.45"S"
            try:
                degrees, minutes, seconds = 0, 0, 0
                is_negative = False
                
                if coord_string.upper().endswith('S') or coord_string.upper().endswith('W'):
                    is_negative = True
                
                clean_str = coord_string.upper().replace('N', '').replace('S', '').replace('E', '').replace('W', '')
                clean_str = clean_str.replace('"', '').replace("'", "'")
                
                # Parse components
                if '°' in clean_str:
                    parts = clean_str.split('°')
                    degrees = float(parts[0])
                    if len(parts) > 1:
                        if "'" in parts[1]:
                            min_sec = parts[1].split("'")
                            minutes = float(min_sec[0])
                            if len(min_sec) > 1 and min_sec[1]:
                                seconds = float(min_sec[1])
                
                # Convert to decimal degrees
                decimal_degrees = degrees + (minutes / 60) + (seconds / 3600)
                
                # Apply the negative sign if needed
                if is_negative:
                    decimal_degrees = -decimal_degrees
                    
                return decimal_degrees
                
            except (ValueError, IndexError, TypeError):
                return pd.NA
        
        # For any other types, try direct conversion
        try:
            return float(coord_value)
        except (ValueError, TypeError):
            return pd.NA

    def process_coordinates(df, lat_col, lon_col, sheet_name=None, excel_name=None): #Process coordinates and error coordinates
        if lat_col is None or lon_col is None:
            print(f"⚠️ Cannot process coordinates in '{sheet_name}': Missing coordinate column(s)")
            return df, []
            
        df_copy = df.copy()
        error_rows = []
        
        # Process each row with better error tracking
        for idx, row in df_copy.iterrows():
            try:
                # Defensive check to make sure the columns exist in the row
                if lat_col not in row or lon_col not in row:
                    print(f"⚠️ Row {idx} missing coordinate column(s) in '{sheet_name}'")
                    continue
                    
                lat_val = parse_coordinate(row[lat_col])
                lon_val = parse_coordinate(row[lon_col])
                
                # Error tracking
                if pd.isna(lat_val) and not pd.isna(row[lat_col]):
                    error_info = {
                        'Excel File': excel_name,
                        'Sheet': sheet_name,
                        'Row Index': idx,
                        'Original Lat Value': row[lat_col],
                        'Original Lon Value': row[lon_col],
                        'Error': f"Failed to parse latitude: {row[lat_col]}"
                    }
                    error_rows.append(error_info)
                
                if pd.isna(lon_val) and not pd.isna(row[lon_col]):
                    error_info = {
                        'Excel File': excel_name,
                        'Sheet': sheet_name,
                        'Row Index': idx,
                        'Original Lat Value': row[lat_col],
                        'Original Lon Value': row[lon_col],
                        'Error': f"Failed to parse longitude: {row[lon_col]}"
                    }
                    error_rows.append(error_info)
                
                df_copy.at[idx, lat_col] = lat_val
                df_copy.at[idx, lon_col] = lon_val
                
            except Exception as e:
                error_info = {
                    'Excel File': excel_name,
                    'Sheet': sheet_name,
                    'Row Index': idx,
                    'Original Lat Value': str(row.get(lat_col, 'Column not found')),
                    'Original Lon Value': str(row.get(lon_col, 'Column not found')),
                    'Error': str(e)
                }
                error_rows.append(error_info)
        
        # Apply fix_coordinates with better error handling
        try:
            fixed_coords = df_copy.apply(fix_coordinates, axis=1, lat_col=lat_col, lon_col=lon_col)
            df_copy[lat_col] = fixed_coords[0]
            df_copy[lon_col] = fixed_coords[1]
        except Exception as e:
            error_info = {
                'Excel File': excel_name,
                'Sheet': sheet_name,
                'Row Index': 'Multiple',
                'Original Lat Value': 'Multiple',
                'Original Lon Value': 'Multiple',
                'Error': f"Batch coordinate fixing failed: {str(e)}"
            }
            error_rows.append(error_info)
        
        return df_copy, error_rows

    def save_to_shapefile(gdf, output_path, batas_wilayah=None, qml_folder=None):  # Save GeoDataFrame to Shapefile
        try:
            gdf = gdf.copy()
            
            # Make sure we have a valid geometry column
            if 'geometry' not in gdf.columns and 'Geometry' not in gdf.columns:
                print(f"❌ Error: No geometry column found in data for {os.path.basename(output_path)}")
                return
                
            # Ensure the GeoDataFrame has a proper geometry column set
            if not isinstance(gdf, gpd.GeoDataFrame):
                if 'geometry' in gdf.columns:
                    gdf = gpd.GeoDataFrame(gdf, geometry='geometry', crs="EPSG:4326")
                elif 'Geometry' in gdf.columns:
                    gdf = gpd.GeoDataFrame(gdf, geometry='Geometry', crs="EPSG:4326")
                else:
                    print(f"❌ Error: Cannot create GeoDataFrame - no geometry column found")
                    return
            else:
                # Explicitly set the geometry column even if it's already a GeoDataFrame
                if 'geometry' in gdf.columns:
                    gdf = gdf.set_geometry('geometry')
                elif 'Geometry' in gdf.columns:
                    gdf = gdf.set_geometry('Geometry')
            
            # Ensure the GeoDataFrame has a valid CRS
            if gdf.crs is None:
                gdf = gdf.set_crs("EPSG:4326")
            
            # Perform spatial join with batas_wilayah if provided
            if batas_wilayah is not None:
                try:                
                    # Make sure we have valid geometries
                    gdf = gdf[~gdf.geometry.isna()].copy()
                    
                    # Make sure CRS matches for join
                    if gdf.crs != batas_wilayah.crs:
                        gdf = gdf.to_crs(batas_wilayah.crs)
                    
                    # Perform the spatial join
                    gdf = gpd.sjoin(gdf, batas_wilayah[['geometry', 'NAMOBJ']], how="left", predicate="intersects")
                    
                    # Clean up index column created by spatial join
                    if 'index_right' in gdf.columns:
                        gdf = gdf.drop(columns=['index_right'])
                        
                except Exception as e:
                    print(f"Warning: Error during spatial join: {str(e)}")
            
            # Truncate column names to 10 characters but preserve 'name' or 'nama' fields
            new_columns = {}
            for col in gdf.columns:
                if col == "geometry" or col == "Geometry":
                    continue
                
                # Special handling for name-related columns - don't add numbers to these
                if 'name' in str(col).lower() or 'nama' in str(col).lower() or col == 'NAMOBJ':
                    # Just truncate to 10 chars without adding a number
                    new_name = str(col)[:10]
                else:
                    # For all other columns, create a unique name that's short enough
                    base_name = str(col)[:7]
                    suffix = 1
                    new_name = base_name
                    
                    # If this name is already used, add numbers until we find a unique one
                    while new_name in new_columns.values() and len(new_name) < 10:
                        new_name = f"{base_name}_{suffix}"[:10]
                        suffix += 1
                
                new_columns[col] = new_name
            
            # Create a mapping file to record the original field names
            mapping_dict = {v: k for k, v in new_columns.items()}
            
            # Rename the columns
            gdf = gdf.rename(columns=new_columns)
            
            # Modify the output path to include NAMOBJ and "Jalan Eksisting"
            if 'NAMOBJ' in gdf.columns:
                # Group by NAMOBJ and save each group to the appropriate directory
                for name_obj, group in gdf.groupby('NAMOBJ'):
                    if pd.isna(name_obj):
                        name_obj = "Unknown"
                        
                    # Create directory structure: output_folder/Extract Shapefile/NAMOBJ/Jalan Eksisting
                    output_dir = os.path.dirname(output_path)
                    file_name = os.path.basename(output_path)
                    
                    # Create new path with NAMOBJ and Jalan Eksisting folders
                    new_output_dir = os.path.join(output_dir, name_obj, "Jalan Eksisting")
                    os.makedirs(new_output_dir, exist_ok=True)
                    
                    new_output_path = os.path.join(new_output_dir, file_name)
                    
                    # Save the shapefile
                    group.to_file(new_output_path, driver="ESRI Shapefile")
                    
                    print(f"✅ Saved: {new_output_path}")
                    
                    # Apply QML Style if qml_folder is provided
                    if qml_folder is not None:
                        sheet_name = os.path.splitext(file_name)[0].split('_')[-1]  # Extract sheet name from filename
                        qml_source_file = os.path.join(qml_folder, f"{sheet_name}.qml")  # Assuming QML file follows sheet name
                        qml_target_file = new_output_path.replace(".shp", ".qml")
                        
                        if os.path.exists(qml_source_file):
                            import shutil
                            shutil.copy(qml_source_file, qml_target_file)
                            print(f"✅ Applied QML style: {qml_target_file}")
                        else:
                            print(f"⚠️ No QML file found for {sheet_name}")
            else:
                # If NAMOBJ is not in columns, just save to the original path
                gdf.to_file(output_path, driver="ESRI Shapefile")
                print(f"✅ Saved: {output_path}")
                
                # Apply QML Style if qml_folder is provided
                if qml_folder is not None:
                    sheet_name = os.path.splitext(os.path.basename(output_path))[0].split('_')[-1]  # Extract sheet name
                    qml_source_file = os.path.join(qml_folder, f"{sheet_name}.qml")  # Assuming QML file follows sheet name
                    qml_target_file = output_path.replace(".shp", ".qml")
                    
                    if os.path.exists(qml_source_file):
                        import shutil
                        shutil.copy(qml_source_file, qml_target_file)
                        print(f"✅ Applied QML style: {qml_target_file}")
                    else:
                        print(f"⚠️ No QML file found for {sheet_name}")
            
        except Exception as e:
            print(f"❌ Error saving shapefile {output_path}: {str(e)}")
            import traceback
            traceback.print_exc()

    def flatten_excel_to_shapefile(file_path, output_folder, excel_name=None, batas_wilayah=None, qml_folder=None, error_logs=None): #Convert an Excel file to Shapefile and collect error logs
        if error_logs is None:
            error_logs = []
        
        try:
            # If excel_name is not provided, extract it from the file_path
            if excel_name is None:
                excel_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # Create output folder if it doesn't exist
            os.makedirs(output_folder, exist_ok=True)
            
            # Load workbook
            wb = load_workbook(file_path, data_only=True)

            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]

                    # Unmerge cells and fill values
                    for merge in list(ws.merged_cells):
                        ws.unmerge_cells(str(merge))
                        top_left = ws.cell(merge.min_row, merge.min_col).value
                        for row in range(merge.min_row, merge.max_row + 1):
                            for col in range(merge.min_col, merge.max_col + 1):
                                ws.cell(row, col, top_left)

                    # Convert to DataFrame
                    data = list(ws.values)
                    df = pd.DataFrame(data)

                    # Check if DataFrame is empty or has too few rows
                    if df.empty or len(df) < 3:
                        print(f"⚠️ Skipping '{sheet_name}' (Empty sheet or insufficient data)")
                        continue

                    # Find rows containing "NO" (safer approach)
                    try:
                        header_indices = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index
                        if len(header_indices) == 0:
                            print(f"⚠️ Skipping '{sheet_name}' (No header row with 'NO' found)")
                            continue
                        header_index = header_indices[0]
                    except Exception as e:
                        print(f"⚠️ Skipping '{sheet_name}' (Error finding header row: {str(e)})")
                        continue

                    # Use the identified header row
                    df.columns = df.iloc[header_index].astype(str).str.strip()

                    # Remove empty columns
                    df = df.dropna(axis=1, how="all")

                    # First pass of REKAP filtering
                    try:
                        rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering REKAP columns - {str(e)}")

                    # Drop rows above and including header, plus the empty row after header
                    rows_to_drop = list(range(0, header_index + 2))
                    if len(df) > max(rows_to_drop) + 1:  # Make sure we have enough rows
                        df = df.drop(index=rows_to_drop).reset_index(drop=True)
                    else:
                        print(f"⚠️ Skipping '{sheet_name}' (Not enough data rows after header)")
                        continue

                    # Improved header row merging
                    if len(df) >= 2:
                        first_row = df.iloc[0].astype(str).replace('None', '').replace('nan', '')
                        second_row = df.iloc[1].astype(str).replace('None', '').replace('nan', '')
                        
                        # Smart merging to avoid duplication
                        merged_header = []
                        for a, b in zip(first_row, second_row):
                            a = a.strip()
                            b = b.strip()
                            
                            # Skip columns with "REKAP" in the name
                            if "rekap" in a.lower() or "rekap" in b.lower():
                                merged_header.append("TO_BE_REMOVED")  # Mark for removal
                                continue
                            
                            if not a and not b:
                                merged_header.append("Column_" + str(len(merged_header)))
                            elif not a:
                                merged_header.append(b)
                            elif not b:
                                merged_header.append(a)
                            else:
                                if a.lower() in b.lower():
                                    merged_header.append(b)
                                elif b.lower() in a.lower():
                                    merged_header.append(a)
                                else:
                                    merged_header.append(f"{a} {b}")
                        
                        # Ensure column names are unique
                        df.columns = unique_column_names(merged_header)
                        
                        # Remove any columns marked for removal
                        df = df.loc[:, ~df.columns.str.contains("TO_BE_REMOVED")]
                        
                        # Remove the first two rows used for headers
                        df = df.drop(index=[0, 1]).reset_index(drop=True)
                    else:
                        print(f"⚠️ Skipping '{sheet_name}' (Not enough rows for headers)")
                        continue

                    # Second pass of REKAP filtering after merging headers
                    try:
                        rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering REKAP columns (second pass) - {str(e)}")

                    # Normalize column names for consistent detection
                    df.columns = [str(col).lower().strip() if col is not None else f"col_{i}" for i, col in enumerate(df.columns)]

                    # Third pass of REKAP filtering after normalizing
                    try:
                        rekap_mask = df.columns.str.contains("rekap", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering rekap columns (third pass) - {str(e)}")

                    # Check if this sheet is about MARKA or PAGAR PENGAMAN
                    is_marka_sheet = "marka" in sheet_name.lower() or any(col for col in df.columns if isinstance(col, str) and "marka" in col.lower())
                    is_pagar_pengaman_sheet = "pagar pengaman" in sheet_name.lower() or any(col for col in df.columns if isinstance(col, str) and "pagar pengaman" in col.lower())

                    # Find all coordinate columns using the improved function
                    start_lat_col = find_coordinate_columns(df, 'start', 'lat')
                    start_lon_col = find_coordinate_columns(df, 'start', 'lon')
                    end_lat_col = find_coordinate_columns(df, 'end', 'lat')
                    end_lon_col = find_coordinate_columns(df, 'end', 'lon')

                    # Check available coordinate patterns
                    has_start_coords = start_lat_col is not None and start_lon_col is not None
                    has_end_coords = end_lat_col is not None and end_lon_col is not None

                    if not has_start_coords:
                        print(f"⚠️ Skipping '{sheet_name}' (No valid start coordinate columns found)")
                        continue
                    
                    # Filter rows where both latitude and longitude values are blank
                    if has_start_coords:
                        # Convert coordinates to appropriate types and handle formatting issues
                        coord_mask = df[start_lat_col].astype(str).str.strip().replace('', np.nan).notna() & \
                                    df[start_lon_col].astype(str).str.strip().replace('', np.nan).notna()
                        
                        if coord_mask.any():
                            # Only process rows with actual coordinate data
                            df_with_coords = df[coord_mask].reset_index(drop=True)
                            df_processed, start_errors = process_coordinates(df_with_coords, start_lat_col, start_lon_col, sheet_name, excel_name)
                            
                            # Update only the rows that had coordinates
                            df = df.copy()
                            df.loc[coord_mask, :] = df_processed
                            
                            error_logs.extend(start_errors)
                        else:
                            print(f"⚠️ No valid start coordinates found in '{sheet_name}'")
                            continue

                    if has_end_coords:
                        # Similar filtering for end coordinates
                        coord_mask = df[end_lat_col].astype(str).str.strip().replace('', np.nan).notna() & \
                                    df[end_lon_col].astype(str).str.strip().replace('', np.nan).notna()
                        
                        if coord_mask.any():
                            df_with_coords = df[coord_mask].reset_index(drop=True)
                            df_processed, end_errors = process_coordinates(df_with_coords, end_lat_col, end_lon_col, sheet_name, excel_name)
                            
                            df.loc[coord_mask, :] = df_processed
                            error_logs.extend(end_errors)

                    # Check if the end coordinates actually contain valid data
                    if has_end_coords:
                        has_valid_end_coords = not df[end_lat_col].isna().all() and not df[end_lon_col].isna().all()
                        valid_pairs = ((df[start_lat_col].notna() & df[start_lon_col].notna()) & (df[end_lat_col].notna() & df[end_lon_col].notna())).any()
                    else:
                        has_valid_end_coords = False
                        valid_pairs = False
                    
                    # Determine geometry type based on available coordinates, actual data, and sheet type
                    # MARKA sheets should use MultiPoint geometry
                    if is_marka_sheet:
                        print(f"Processing '{sheet_name}' as MultiPoint (MARKA sheet)")

                        # First check if we have any valid coordinates before applying
                        has_valid_coords = ((df[start_lat_col].notna() & df[start_lon_col].notna()) | 
                                            (has_end_coords and df[end_lat_col].notna() & df[end_lon_col].notna())).any()
                        
                        if not has_valid_coords:
                            print(f"⚠️ Skipping '{sheet_name}' (No valid coordinates found in MARKA sheet)")
                            continue
                            
                        # Create MultiPoint geometry only for rows with valid coordinates
                        def create_multipoint(row):
                            try:
                                points = []
                                # Add start point if valid
                                if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]):
                                    points.append((float(row[start_lon_col]), float(row[start_lat_col])))
                                    
                                # Add end point if valid and available
                                if has_end_coords and pd.notna(row[end_lon_col]) and pd.notna(row[end_lat_col]):
                                    points.append((float(row[end_lon_col]), float(row[end_lat_col])))
                                    
                                # Return MultiPoint if we have points, else None
                                return MultiPoint(points) if points else None
                            except Exception as e:
                                print(f"Error creating MultiPoint for row {row.name}: {str(e)}")
                                return None
                        
                        # Apply the function to create geometry
                        df["geometry"] = df.apply(create_multipoint, axis=1)
                        
                        # Exclude coordinate columns from properties
                        exclude_cols = [start_lat_col, start_lon_col]
                        if has_end_coords:
                            exclude_cols.extend([end_lat_col, end_lon_col])
                        exclude_cols.append("geometry")
                        
                    # PAGAR PENGAMAN sheets with valid start/end coordinates should use LineString
                    elif is_pagar_pengaman_sheet and has_valid_end_coords and valid_pairs:
                        print(f"Processing '{sheet_name}' as LineString (PAGAR PENGAMAN sheet)")
                        
                        # Create LineString geometry with additional debugging and more flexible validation
                        def create_linestring(row):
                            try:
                                # Enhanced debugging
                                start_lat = row[start_lat_col]
                                start_lon = row[start_lon_col]
                                end_lat = row[end_lat_col]
                                end_lon = row[end_lon_col]
                                
                                # Skip processing if all coordinates are blank/NaN
                                if pd.isna(start_lat) and pd.isna(start_lon) and pd.isna(end_lat) and pd.isna(end_lon):
                                    return None
                                    
                                # First try to parse any string coordinates
                                if isinstance(start_lat, str):
                                    start_lat = parse_coordinate(start_lat)
                                if isinstance(start_lon, str):
                                    start_lon = parse_coordinate(start_lon)
                                if isinstance(end_lat, str):
                                    end_lat = parse_coordinate(end_lat)
                                if isinstance(end_lon, str):
                                    end_lon = parse_coordinate(end_lon)
                                
                                # Create LineString if we have valid coordinates
                                if (pd.notna(start_lon) and pd.notna(start_lat) and 
                                    pd.notna(end_lon) and pd.notna(end_lat)):
                                    return LineString([
                                        (float(start_lon), float(start_lat)),
                                        (float(end_lon), float(end_lat))
                                    ])
                                else:
                                    # If one set of coordinates is missing, fall back to Point
                                    if pd.notna(start_lon) and pd.notna(start_lat):
                                        offset = 0.0001
                                        return LineString([
                                            (float(start_lon), float(start_lat)),
                                            (float(start_lon) + offset, float(start_lat) + offset)
                                        ])
                                    return None
                            except Exception as e:
                                print(f"Error creating LineString for row {row.name}: {str(e)}")
                                return None
                        
                        df["geometry"] = df.apply(create_linestring, axis=1)
                        exclude_cols = [start_lat_col, start_lon_col, end_lat_col, end_lon_col, "geometry"]
                        
                    else:
                        # Other sheets use Point geometry (only start coordinates)
                        print(f"Processing '{sheet_name}' as Point geometry (regular sheet)")
                        
                        # Create Point geometry with start coordinates, only for rows with valid data
                        df["geometry"] = df.apply(
                            lambda row: Point(row[start_lon_col], row[start_lat_col]) 
                            if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) else None,
                            axis=1
                        )
                        
                        # Exclude coordinate columns from properties
                        exclude_cols = [start_lat_col, start_lon_col]
                        if has_end_coords:
                            exclude_cols.extend([end_lat_col, end_lon_col])
                        exclude_cols.append("geometry")

                    # Drop rows where geometry is None
                    df = df.dropna(subset=["geometry"]).reset_index(drop=True)

                    # Only create GeoDataFrame if there are valid geometries
                    if len(df) > 0 and not df["geometry"].isnull().all():
                        # Filter REKAP columns before creating properties
                        properties_cols = [col for col in df.columns if col not in exclude_cols]
                        properties_cols = [col for col in properties_cols if 'rekap' not in str(col).lower()]
                        
                        # Create GeoDataFrame with explicit geometry setting
                        gdf = gpd.GeoDataFrame(df[properties_cols + ["geometry"]], geometry="geometry",crs="EPSG:4326")
                        
                        # Verify that the geometry column is properly set
                        if not gdf.geometry.name == "geometry":
                            gdf = gdf.set_geometry("geometry")
                        
                        gdf.columns = clean_column_names(gdf.columns)

                        gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
                        
                        gdf = gdf.loc[:, ~gdf.columns.astype(str).str.match(r"^None$|None_", na=False)]

                        gdf.columns = gdf.columns.astype(str).str.replace(r"\sNone\b", "", regex=True).str.strip()

                        # One final check for REKAP columns before saving
                        if any('rekap' in str(col).lower() for col in gdf.columns):
                            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
                        
                        # Define output file path
                        output_path = os.path.join(output_folder, f"{excel_name}_{sheet_name}.shp")
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)

                        # Save as shapefile
                        save_to_shapefile(gdf, output_path, batas_wilayah, qml_folder)
                    else:
                        print(f"⚠️ Skipping '{sheet_name}' (No valid geometry found)")
                        continue
                except Exception as e:
                    print(f"❌ Error processing sheet '{sheet_name}': {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue

            return error_logs
        except Exception as e:
            print(f"❌ Error processing file: {str(e)}")
            import traceback
            traceback.print_exc()
            return error_logs

    def process_single_excel_file_shapefile(file_path, output_base_folder, qml_folder=None, batas_wilayah_path=None): #Process conversion for one Excel file
        output_folder = os.path.join(output_base_folder, "Extract Shapefile")
        os.makedirs(output_folder, exist_ok=True)
        
        # Load the city boundaries shapefile if provided
        batas_wilayah = None
        if batas_wilayah_path and os.path.exists(batas_wilayah_path):
            try:
                batas_wilayah = gpd.read_file(batas_wilayah_path)
                print(f"✅ Loaded city boundaries from: {batas_wilayah_path}")
            except Exception as e:
                print(f"❌ Error loading city boundaries shapefile: {str(e)}")
        else:
            print("❌ No city boundaries provided or file not found.")
        
        # Check if QML folder exists
        if qml_folder and os.path.exists(qml_folder):
            print(f"✅ Using QML styles from: {qml_folder}")
        else:
            print("❌ No QML folder provided or folder not found.")
            qml_folder = None
        
        # Extract file name
        file_name = os.path.basename(file_path)
        excel_name = os.path.splitext(file_name)[0]
        
        print(f"Processing: {file_name}")
        
        try:
            # Initialize error_logs list
            error_logs = []
            
            wb = load_workbook(file_path, data_only=True)
            
            # Process the file (this will handle all sheets)
            error_logs = flatten_excel_to_shapefile(file_path, output_folder, excel_name, batas_wilayah, qml_folder, error_logs)
            
            print(f"✅ Completed processing: {file_name}")
            if error_logs:
                log_coordinate_errors(error_logs, output_base_folder)
        except Exception as e:
            print(f"❌ Error processing {file_name}: {str(e)}")
            import traceback
            traceback.print_exc()

    def process_excel_folder_shapefile(input_folder, output_base_folder, qml_folder=None, batas_wilayah_path=None): #Process conversion for all Excel files in a folder
        output_folder = os.path.join(output_base_folder, "Extract Shapefile")
        os.makedirs(output_folder, exist_ok=True)
        
        # Load the city boundaries shapefile if provided
        batas_wilayah = None
        if batas_wilayah_path and os.path.exists(batas_wilayah_path):
            try:
                batas_wilayah = gpd.read_file(batas_wilayah_path)
                print(f"✅ Loaded city boundaries from: {batas_wilayah_path}")
            except Exception as e:
                print(f"❌ Error loading city boundaries shapefile: {str(e)}")
        else:
            print("❌ No city boundaries provided or file not found.")
        
        # Check if QML folder exists
        if qml_folder and os.path.exists(qml_folder):
            print(f"✅ Using QML styles from: {qml_folder}")
        else:
            print("❌ No QML folder provided or folder not found.")
            qml_folder = None
        
        # Get all Excel files in the input folder
        excel_extensions = ['*.xlsx', '*.xls', '*.xlsm']
        excel_files = []
        
        for ext in excel_extensions:
            excel_files.extend(glob.glob(os.path.join(input_folder, ext)))
        
        # Initialize error log list
        all_error_logs = []
        
        # Process each Excel file
        for i, file_path in enumerate(excel_files, 1):
            file_name = os.path.basename(file_path)
            excel_name = os.path.splitext(file_name)[0]
            print(f"\n[{i}/{len(excel_files)}] Processing: {file_name}")
            
            try:
                # Collect errors from processing this file
                file_errors = flatten_excel_to_shapefile(file_path, output_folder, excel_name, batas_wilayah, qml_folder, [])
                all_error_logs.extend(file_errors)
                print(f"✅ Completed processing: {file_name}")
                if file_errors:
                    print(f"⚠️ Found {len(file_errors)} coordinate errors during processing")
            except Exception as e:
                print(f"❌ Error processing {file_name}: {str(e)}")
        
        # Log all errors
        if all_error_logs:
            log_coordinate_errors(all_error_logs, output_base_folder)
        
        print(f"\n🎉 All Excel files processed. Output saved to: {output_folder}")

    def extract_images_from_excel(file_path, output_folder): #Extract images from columns containing specific keywords in all sheets and save them.

        # Keywords to look for in column names
        KEYWORDS = ["DOKUMENTASI", "RAMBU", "RPPJ"]
        
        # Extract filename from path for folder creation
        file_name_clean = re.search(r'([^\\]+)\.xlsx$', file_path)
        if file_name_clean:
            file_name_clean = file_name_clean.group(1)
        else:
            file_name_clean = os.path.basename(file_path).replace('.xlsx', '')
        
        # Create subdirectories for each category
        dokumentasi_folder = os.path.join(output_folder, "Dokumentasi", file_name_clean)  # Changed - nested under Excel name
        rambu_folder = os.path.join(output_folder, "Rambu")
        rppj_folder = os.path.join(output_folder, "RPPJ")
        
        for folder in [dokumentasi_folder, rambu_folder, rppj_folder]:
            os.makedirs(folder, exist_ok=True)
        
        try:
            # Process each sheet independently to prevent file handle issues
            wb = load_workbook(file_path, data_only=True)
            sheet_names = wb.sheetnames.copy()
            wb.close()
            
            successful_sheets = 0
            total_images_saved = 0
            
            # Process sheets in the order they appear in the workbook
            for sheet_idx, sheet_name in enumerate(sheet_names, 1):
                try:
                    print(f"Processing sheet {sheet_idx}/{len(sheet_names)}: {sheet_name}")
                    
                    # Create a safe sheet name for filenames
                    safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
                    
                    # Open a fresh workbook instance for each sheet
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb[sheet_name]
                    
                    # Create a fresh image loader for this sheet
                    image_loader = SheetImageLoader(ws)
                    
                    # Step 1: Extract merged column headers from rows 1-5
                    column_names = {}
                    
                    # Get values for rows 1-5 for each column
                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        header_values = []
                        
                        for row in range(1, 6):  # Rows 1-5
                            cell_value = ws[f"{col_letter}{row}"].value
                            if cell_value:
                                header_values.append(str(cell_value).strip())
                        
                        # Combine the header parts into one name
                        if header_values:
                            column_names[col] = " ".join(header_values)
                        else:
                            column_names[col] = f"Column_{col_letter}"
                    
                    # Step 2: Find columns containing our keywords
                    dokumentasi_columns = {}
                    rambu_columns = {}
                    rppj_columns = {}
                    nama_rambu_column = None
                    jenis_tiang_column = None
                    
                    # First, find all our target columns
                    for col, name in column_names.items():
                        if "DOKUMENTASI" in name.upper():
                            dokumentasi_columns[col] = name
                        elif "RAMBU" in name.upper() and "NAMA RAMBU" not in name.upper():
                            rambu_columns[col] = name
                        elif "RPPJ" in name.upper():
                            rppj_columns[col] = name
                        
                        # Find the specific column for "Nama Rambu"
                        if "NAMA RAMBU" in name.upper():
                            nama_rambu_column = col
                            print(f"Found 'Nama Rambu' column: {name} (Column {get_column_letter(col)})")
                        
                        # Find the specific column for "Jenis Tiang"
                        elif "JENIS TIANG" in name.upper():
                            jenis_tiang_column = col
                            print(f"Found 'Jenis Tiang' column: {name} (Column {get_column_letter(col)})")
                    
                    # Track images processed for each category
                    images_by_category = {
                        "dokumentasi": 0,
                        "rambu": 0,
                        "rppj": 0
                    }
                    
                    # Process each category separately
                    # 1. Process DOKUMENTASI columns
                    if dokumentasi_columns:
                        processed = process_image_columns(
                            ws, image_loader, dokumentasi_columns, dokumentasi_folder, 
                            file_name_clean, safe_sheet_name, "dokumentasi", None, None
                        )
                        images_by_category["dokumentasi"] += processed
                    
                    # 2. Process RAMBU columns (custom naming based on "Nama Rambu" column)
                    if rambu_columns:
                        if nama_rambu_column is None:
                            print("⚠️ Warning: 'Nama Rambu' column not found. Using default naming for Rambu images.")
                        
                        processed = process_image_columns(
                            ws, image_loader, rambu_columns, rambu_folder, 
                            file_name_clean, safe_sheet_name, "rambu", nama_rambu_column, None
                        )
                        images_by_category["rambu"] += processed
                    
                    # 3. Process RPPJ columns (custom naming based on "Jenis Tiang" column)
                    if rppj_columns:
                        if jenis_tiang_column is None:
                            print("⚠️ Warning: 'Jenis Tiang' column not found. Using default naming for RPPJ images.")
                        
                        processed = process_image_columns(
                            ws, image_loader, rppj_columns, rppj_folder, 
                            file_name_clean, safe_sheet_name, "rppj", None, jenis_tiang_column
                        )
                        images_by_category["rppj"] += processed
                    
                    total_images_saved += sum(images_by_category.values())
                    
                    print(f"✅ Completed sheet '{sheet_name}':")
                    print(f"  - Dokumentasi: {images_by_category['dokumentasi']} images")
                    print(f"  - Rambu: {images_by_category['rambu']} images")
                    print(f"  - RPPJ: {images_by_category['rppj']} images")
                    
                    if sum(images_by_category.values()) > 0:
                        successful_sheets += 1
                    
                except Exception as e:
                    print(f"❌ Error processing sheet '{sheet_name}' in file '{file_name_clean}': {str(e)}")
                
                finally:
                    # Always close the workbook after processing each sheet
                    if 'wb' in locals() and wb is not None:
                        try:
                            wb.close()
                        except:
                            pass
            
            print(f"✅ Completed processing file: {file_name_clean}")
            print(f"  - {successful_sheets}/{len(sheet_names)} sheets processed")
            print(f"  - {total_images_saved} total images extracted")
            return True
            
        except Exception as e:
            # Handle any errors in the outer scope
            print(f"❌ Error processing file '{file_path}': {str(e)}")
            return False

    def process_image_columns(ws, image_loader, target_columns, output_folder, file_name_clean, #Process images in specified columns with custom naming logic.
                            safe_sheet_name, category, nama_rambu_column, jenis_tiang_column): 
        
        # Track existing image names (to avoid duplicates for Rambu)
        existing_images = {}
        
        # Find image cells in target columns
        image_cells = {}
        for col in target_columns:
            col_letter = get_column_letter(col)
            column_name = target_columns[col]
            
            # Make column name safe for filename
            safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
            
            # Scan rows starting from row 6
            for row in range(6, ws.max_row + 1):
                cell_address = f"{col_letter}{row}"
                if image_loader.image_in(cell_address):
                    image_cells[(row, col)] = {
                        'cell_address': cell_address,
                        'column_name': safe_column_name,
                        'column_letter': col_letter,
                        'row_number': row
                    }
        
        if not image_cells:
            print(f"⚠️ No images found in {category.upper()} columns, skipping...")
            return 0
        
        print(f"Found {len(image_cells)} images in {category.upper()} columns")
        
        # Process images in column-first, row-second order
        successful_images = 0
        image_cells_by_column = {}
        
        # Group by column
        for (row, col), cell_info in image_cells.items():
            if col not in image_cells_by_column:
                image_cells_by_column[col] = []
            image_cells_by_column[col].append((row, cell_info))
        
        # Process each column
        for col in sorted(image_cells_by_column.keys()):
            column_letter = get_column_letter(col)
            column_name = target_columns[col]
            safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
            
            print(f"Processing {len(image_cells_by_column[col])} images in column '{column_name}'")
            
            # Process rows in order
            for row, cell_info in sorted(image_cells_by_column[col], key=lambda x: x[0]):
                cell_address = cell_info['cell_address']
                try:
                    # Get the image
                    img = image_loader.get(cell_address)
                    
                    # Generate filename based on category
                    if category == "rambu" and nama_rambu_column is not None:
                        # For RAMBU: Use "Nama Rambu" column value as filename
                        nama_rambu_value = ws[f"{get_column_letter(nama_rambu_column)}{row}"].value
                        
                        if nama_rambu_value and str(nama_rambu_value).strip():
                            # Use the actual value from "Nama Rambu" column
                            safe_nama_rambu = re.sub(r'[\\/*?:"<>|]', "_", str(nama_rambu_value).strip())
                            img_filename = f"{safe_nama_rambu}.png"
                            
                            # Check if this name already exists (to avoid duplicates)
                            if img_filename in existing_images:
                                print(f"  ⚠️ Duplicate 'Nama Rambu' found: {safe_nama_rambu} - Replacing existing image")
                            
                            existing_images[img_filename] = True
                        else:
                            # Fallback to default naming if no nama_rambu value
                            print(f"  ⚠️ No 'Nama Rambu' value found for row {row}, using default naming")
                            row_identifier = f"Row{row}"
                            safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                            img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                    
                    elif category == "rppj" and jenis_tiang_column is not None:
                        # For RPPJ: Use "Jenis Tiang" column value in filename
                        jenis_tiang_value = ws[f"{get_column_letter(jenis_tiang_column)}{row}"].value
                        safe_jenis_tiang = "Unknown"
                        if jenis_tiang_value and str(jenis_tiang_value).strip():
                            safe_jenis_tiang = re.sub(r'[\\/*?:"<>|]', "_", str(jenis_tiang_value).strip())
                        
                        # Get row identifier - ALWAYS use just the row number without column A value
                        row_identifier = f"Row{row}"
                        
                        safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                        img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_jenis_tiang}_{safe_row_identifier}.png"
                    
                    else:
                        # Default naming for DOKUMENTASI (and fallback for others)
                        # ALWAYS use just the row number without column A value
                        row_identifier = f"Row{row}"
                        
                        safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                        img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                    
                    # Save the image
                    img_path = os.path.join(output_folder, img_filename)
                    with io.BytesIO() as img_buffer:
                        img.save(img_buffer, format="PNG")
                        img_buffer.seek(0)
                        with open(img_path, 'wb') as f:
                            f.write(img_buffer.read())
                    
                    successful_images += 1
                    print(f"  ✅ Saved: {img_filename}")
                except Exception as e:
                    print(f"  ❌ Error saving image at {cell_address}: {str(e)}")
        
        return successful_images

    def process_single_excel_file_images(file_path, export_folder): #Process a single Excel file and extract images from it.
        # Create "Extract Images" folder within the export directory
        output_folder = os.path.join(export_folder, "Extract Images")
        os.makedirs(output_folder, exist_ok=True)
        
        # Process the file
        file_name = os.path.basename(file_path)
        print(f"\n📊 Processing file: {file_name}")
        
        result = extract_images_from_excel(file_path, output_folder)
        
        # Print summary
        print("\n" + "="*50)
        print("📈 PROCESSING SUMMARY")
        print("="*50)
        print(f"File: {file_name}")
        
        if result:
            print(f"Status: Successfully processed ✅")
        else:
            print(f"Status: Failed to process ❌")
        
        print(f"Images saved to: {output_folder}")
        print("\n🎉 Processing completed!")
        
        return result

    def process_excel_folder_images(folder_path, export_folder): #Process all Excel files in a folder and extract images from them.

        # Create "Extract Images" folder within the export directory
        output_folder = os.path.join(export_folder, "Extract Images")
        os.makedirs(output_folder, exist_ok=True)
        
        # Track statistics
        total_files = 0
        successful_files = 0
        failed_files = []
        
        # Get all Excel files in the folder - with full paths
        excel_files = []
        for file in os.listdir(folder_path):
            if file.endswith(('.xlsx', '.xlsm')):
                excel_files.append(os.path.join(folder_path, file))
        
        if not excel_files:
            print("⚠️ No Excel files found in the specified folder.")
            return
        
        total_files = len(excel_files)  # Fixed variable assignment
        
        # Process each Excel file
        for i, file_path in enumerate(excel_files, 1):
            file_name = os.path.basename(file_path)
            print(f"\n📊 Processing file {i}/{total_files}: {file_name}")
            
            if extract_images_from_excel(file_path, output_folder):
                successful_files += 1
            else:
                failed_files.append(file_name)
        
        # Print summary
        print("\n" + "="*50)
        print("📈 PROCESSING SUMMARY")
        print("="*50)
        print(f"Total files: {total_files}")
        print(f"Successfully processed: {successful_files}")
        print(f"Failed to process: {len(failed_files)}")
        print(f"Images saved to: {output_folder}")
        
        if failed_files:
            print("\nFiles that could not be processed:")
            for file in failed_files:
                print(f"- {file}")
        
        print("\n🎉 All Excel files processing completed!")

    def sanitize_for_path(text):
        if text is None:
            return "unknown"
        # Replace characters that are not safe for file paths
        unsafe_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        result = str(text)
        for char in unsafe_chars:
            result = result.replace(char, '_')
        return result.strip()

    def add_image_documentation_paths(gdf, excel_name, sheet_name, output_base_dir):
        # Find documentation column
        doc_columns = [col for col in gdf.columns if 'dokumentasi' in str(col).lower()]
        
        if not doc_columns:
            # If no dokumentasi column exists, add a placeholder one
            gdf['Image Dokumentasi'] = None
            doc_columns = ['Image Dokumentasi']
        
        # Use the first found dokumentasi column
        doc_column = doc_columns[0]
        
        # Clean names for use in paths
        file_name_clean = sanitize_for_path(excel_name)
        safe_sheet_name = sanitize_for_path(sheet_name)
        safe_column_name = sanitize_for_path(doc_column)
        
        # Generate image paths for each row
        for idx, row in gdf.iterrows():
            safe_row_identifier = sanitize_for_path(idx)
            
            # Get Kota/Kabupaten name, if available
            kota_kab = "Unknown"
            if 'Kota/Kabupaten' in gdf.columns and pd.notna(row['Kota/Kabupaten']):
                kota_kab = sanitize_for_path(row['Kota/Kabupaten'])
            
            # Construct the image path with the requested structure
            image_path = os.path.join(
                output_base_dir,
                "Extract Images",
                "Dokumentasi",
                f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_Row{safe_row_identifier}.png"
            )
            
            # Assign to the dataframe
            gdf.at[idx, 'Image Dokumentasi'] = image_path
        
        return gdf

    def add_image_paths(gdf, excel_name, sheet_name, output_base_dir):
        # Clean names for use in paths
        file_name_clean = sanitize_for_path(excel_name)
        safe_sheet_name = sanitize_for_path(sheet_name)
        
        # Check if it's a Rambu sheet
        if 'rambu' in sheet_name.lower():
            # Check if 'Nama Rambu' column exists
            nama_rambu_columns = [col for col in gdf.columns if 'nama rambu' in str(col).lower()]
            
            if nama_rambu_columns:
                nama_rambu_column = nama_rambu_columns[0]
                
                # Generate image paths for each row
                for idx, row in gdf.iterrows():
                    nama_rambu_value = str(row.get(nama_rambu_column, ''))
                    
                    if nama_rambu_value and nama_rambu_value.lower() != 'nan' and nama_rambu_value.lower() != 'none':
                        # Construct the image path using the Nama Rambu value
                        safe_nama_rambu = sanitize_for_path(nama_rambu_value)
                        image_path = os.path.join(output_base_dir, "Extract Images", "Rambu", f"{safe_nama_rambu}.png")
                        
                        # Assign the constructed path to the 'Image Rambu' property
                        gdf.at[idx, 'Image Rambu'] = image_path
                    else:
                        gdf.at[idx, 'Image Rambu'] = None  # Handle missing Nama Rambu values
        
        # Check if it's an RPPJ sheet
        elif 'rppj' in sheet_name.lower():
            # Generate image paths for each row
            for idx, row in gdf.iterrows():
                safe_row_identifier = sanitize_for_path(idx)
                
                # Construct the image path for RPPJ
                image_path = os.path.join(
                    output_base_dir,
                    "Extract Images",
                    "RPPJ",
                    f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_RPPJ_Row{safe_row_identifier}.png"
                )
                
                # Assign to the dataframe
                gdf.at[idx, 'Image RPPJ'] = image_path
        
        return gdf

    def save_to_geojson(gdf, output_path, batas_wilayah=None, excel_name=None, sheet_name=None, output_base_dir=None):
        try:
            gdf = gdf.copy()
            
            # Make sure we have a valid geometry column
            if 'geometry' not in gdf.columns and 'Geometry' not in gdf.columns:
                print(f"❌ Error: No geometry column found in data for {os.path.basename(output_path)}")
                return
                
            # Ensure the GeoDataFrame has a proper geometry column set
            if not isinstance(gdf, gpd.GeoDataFrame):
                if 'geometry' in gdf.columns:
                    gdf = gpd.GeoDataFrame(gdf, geometry='geometry', crs="EPSG:4326")
                elif 'Geometry' in gdf.columns:
                    gdf = gpd.GeoDataFrame(gdf, geometry='Geometry', crs="EPSG:4326")
                else:
                    print(f"❌ Error: Cannot create GeoDataFrame - no geometry column found")
                    return
            else:
                # Explicitly set the geometry column even if it's already a GeoDataFrame
                if 'geometry' in gdf.columns:
                    gdf = gdf.set_geometry('geometry')
                elif 'Geometry' in gdf.columns:
                    gdf = gdf.set_geometry('Geometry')
            
            # Ensure the GeoDataFrame has a valid CRS
            if gdf.crs is None:
                gdf = gdf.set_crs("EPSG:4326")
            
            # Perform spatial join with batas_wilayah if provided
            if batas_wilayah is not None:
                try:                
                    # Make sure we have valid geometries
                    gdf = gdf[~gdf.geometry.isna()].copy()
                    
                    # Make sure CRS matches for join
                    if gdf.crs != batas_wilayah.crs:
                        gdf = gdf.to_crs(batas_wilayah.crs)
                    
                    # Perform the spatial join
                    gdf = gpd.sjoin(gdf, batas_wilayah[['geometry', 'NAMOBJ']], how="left", predicate="intersects")
                    
                    # Rename NAMOBJ column to Kota/Kabupaten
                    if 'NAMOBJ' in gdf.columns:
                        gdf = gdf.rename(columns={'NAMOBJ': 'Kota/Kabupaten'})
                    
                    # Clean up index column created by spatial join
                    if 'index_right' in gdf.columns:
                        gdf = gdf.drop(columns=['index_right'])
                        
                except Exception as e:
                    print(f"Warning: Error during spatial join: {str(e)}")
            
            # Add the image documentation paths before saving
            if excel_name is not None and sheet_name is not None and output_base_dir is not None:
                # Add documentation image paths
                gdf = add_image_documentation_paths(gdf, excel_name, sheet_name, output_base_dir)
                
                # Add new special image paths based on sheet type
                gdf = add_image_paths(gdf, excel_name, sheet_name, output_base_dir)
            
            # Modify the output path to include Kota/Kabupaten and "Jalan Eksisting"
            if 'Kota/Kabupaten' in gdf.columns:
                # Group by Kota/Kabupaten and save each group to the appropriate directory
                for name_obj, group in gdf.groupby('Kota/Kabupaten'):
                    if pd.isna(name_obj):
                        name_obj = "Unknown"
                        
                    # Create directory structure: output_folder/Extract GeoJSON/Kota/Kabupaten/Jalan Eksisting
                    output_dir = os.path.dirname(output_path)
                    file_name = os.path.basename(output_path)
                    
                    # Create new path with Kota/Kabupaten and Jalan Eksisting folders
                    new_output_dir = os.path.join(output_dir, name_obj, "Jalan Eksisting")
                    os.makedirs(new_output_dir, exist_ok=True)
                    
                    new_output_path = os.path.join(new_output_dir, file_name)
                    
                    # Save the GeoJSON file
                    clean_geojson(group, new_output_path)
            else:
                # If Kota/Kabupaten is not in columns, just save to the original path
                clean_geojson(gdf, output_path)
        except Exception as e:
            print(f"❌ Error saving GeoJSON {output_path}: {str(e)}")
            import traceback
            traceback.print_exc()

    def clean_geojson(gdf, output_path):  # Save GeoDataFrame in a clean format GeoJSON file
        
        temp_path = output_path.replace(".geojson", "_temp.geojson")
        gdf.to_file(temp_path, driver="GeoJSON")
        
        with open(temp_path, "r", encoding="utf-8") as file:
            geojson_data = json.load(file)
        
        with open(output_path, "w", encoding="utf-8") as file:
            json.dump(geojson_data, file, indent=4)
        
        os.remove(temp_path)
        print(f"✅ Saved: {output_path}")

    def flatten_excel_to_geojson(file_path, output_folder, excel_name=None, batas_wilayah=None, error_logs=None):
        if error_logs is None:
            error_logs = []
        
        try:
            # If excel_name is not provided, extract it from the file_path
            if excel_name is None:
                excel_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # Create output folder if it doesn't exist
            os.makedirs(output_folder, exist_ok=True)
            
            # Extract the output base directory from output_folder
            output_base_dir = os.path.dirname(os.path.dirname(output_folder))
            
            # Load workbook
            wb = load_workbook(file_path, data_only=True)

            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]

                    # Unmerge cells and fill values
                    for merge in list(ws.merged_cells):
                        ws.unmerge_cells(str(merge))
                        top_left = ws.cell(merge.min_row, merge.min_col).value
                        for row in range(merge.min_row, merge.max_row + 1):
                            for col in range(merge.min_col, merge.max_col + 1):
                                ws.cell(row, col, top_left)

                    # Convert to DataFrame
                    data = list(ws.values)
                    df = pd.DataFrame(data)

                    # Check if DataFrame is empty or has too few rows
                    if df.empty or len(df) < 3:
                        print(f"⚠️ Skipping '{sheet_name}' (Empty sheet or insufficient data)")
                        continue

                    # Find rows containing "NO" (safer approach)
                    try:
                        header_indices = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index
                        if len(header_indices) == 0:
                            print(f"⚠️ Skipping '{sheet_name}' (No header row with 'NO' found)")
                            continue
                        header_index = header_indices[0]
                    except Exception as e:
                        print(f"⚠️ Skipping '{sheet_name}' (Error finding header row: {str(e)})")
                        continue

                    # Use the identified header row
                    df.columns = df.iloc[header_index].astype(str).str.strip()

                    # Remove empty columns
                    df = df.dropna(axis=1, how="all")

                    # First pass of REKAP filtering
                    try:
                        rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering REKAP columns - {str(e)}")

                    # Drop rows above and including header, plus the empty row after header
                    rows_to_drop = list(range(0, header_index + 2))
                    if len(df) > max(rows_to_drop) + 1:  # Make sure we have enough rows
                        df = df.drop(index=rows_to_drop).reset_index(drop=True)
                    else:
                        print(f"⚠️ Skipping '{sheet_name}' (Not enough data rows after header)")
                        continue

                    # Improved header row merging
                    if len(df) >= 2:
                        first_row = df.iloc[0].astype(str).replace('None', '').replace('nan', '')
                        second_row = df.iloc[1].astype(str).replace('None', '').replace('nan', '')
                        
                        # Smart merging to avoid duplication
                        merged_header = []
                        for a, b in zip(first_row, second_row):
                            a = a.strip()
                            b = b.strip()
                            
                            # Skip columns with "REKAP" in the name
                            if "rekap" in a.lower() or "rekap" in b.lower():
                                merged_header.append("TO_BE_REMOVED")  # Mark for removal
                                continue
                            
                            if not a and not b:
                                merged_header.append("Column_" + str(len(merged_header)))
                            elif not a:
                                merged_header.append(b)
                            elif not b:
                                merged_header.append(a)
                            else:
                                if a.lower() in b.lower():
                                    merged_header.append(b)
                                elif b.lower() in a.lower():
                                    merged_header.append(a)
                                else:
                                    merged_header.append(f"{a} {b}")
                        
                        # Ensure column names are unique
                        df.columns = unique_column_names(merged_header)
                        
                        # Remove any columns marked for removal
                        df = df.loc[:, ~df.columns.str.contains("TO_BE_REMOVED")]
                        
                        # Remove the first two rows used for headers
                        df = df.drop(index=[0, 1]).reset_index(drop=True)
                    else:
                        print(f"⚠️ Skipping '{sheet_name}' (Not enough rows for headers)")
                        continue

                    # Second pass of REKAP filtering after merging headers
                    try:
                        rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering REKAP columns (second pass) - {str(e)}")

                    # Normalize column names for consistent detection
                    df.columns = [str(col).lower().strip() if col is not None else f"col_{i}" for i, col in enumerate(df.columns)]

                    # Third pass of REKAP filtering after normalizing
                    try:
                        rekap_mask = df.columns.str.contains("rekap", case=False, na=False)
                        if rekap_mask.any():
                            df = df.loc[:, ~rekap_mask]
                    except Exception as e:
                        print(f"Warning in '{sheet_name}': Error filtering rekap columns (third pass) - {str(e)}")

                    # Check if this sheet is about MARKA or PAGAR PENGAMAN
                    is_marka_sheet = "marka" in sheet_name.lower() or any(col for col in df.columns if isinstance(col, str) and "marka" in col.lower())
                    is_pagar_pengaman_sheet = "pagar pengaman" in sheet_name.lower() or any(col for col in df.columns if isinstance(col, str) and "pagar pengaman" in col.lower())

                    # Find all coordinate columns using the improved function
                    start_lat_col = find_coordinate_columns(df, 'start', 'lat')
                    start_lon_col = find_coordinate_columns(df, 'start', 'lon')
                    end_lat_col = find_coordinate_columns(df, 'end', 'lat')
                    end_lon_col = find_coordinate_columns(df, 'end', 'lon')

                    # Check available coordinate patterns
                    has_start_coords = start_lat_col is not None and start_lon_col is not None
                    has_end_coords = end_lat_col is not None and end_lon_col is not None

                    if not has_start_coords:
                        print(f"⚠️ Skipping '{sheet_name}' (No valid start coordinate columns found)")
                        continue
                    
                    # Filter rows where both latitude and longitude values are blank
                    if has_start_coords:
                        # Convert coordinates to appropriate types and handle formatting issues
                        coord_mask = df[start_lat_col].astype(str).str.strip().replace('', np.nan).notna() & \
                                    df[start_lon_col].astype(str).str.strip().replace('', np.nan).notna()
                        
                        if coord_mask.any():
                            # Only process rows with actual coordinate data
                            df_with_coords = df[coord_mask].reset_index(drop=True)
                            df_processed, start_errors = process_coordinates(df_with_coords, start_lat_col, start_lon_col, sheet_name, excel_name)
                            
                            # Update only the rows that had coordinates
                            df = df.copy()
                            df.loc[coord_mask, :] = df_processed
                            
                            error_logs.extend(start_errors)
                        else:
                            print(f"⚠️ No valid start coordinates found in '{sheet_name}'")
                            continue

                    if has_end_coords:
                        # Similar filtering for end coordinates
                        coord_mask = df[end_lat_col].astype(str).str.strip().replace('', np.nan).notna() & \
                                    df[end_lon_col].astype(str).str.strip().replace('', np.nan).notna()
                        
                        if coord_mask.any():
                            df_with_coords = df[coord_mask].reset_index(drop=True)
                            df_processed, end_errors = process_coordinates(df_with_coords, end_lat_col, end_lon_col, sheet_name, excel_name)
                            
                            df.loc[coord_mask, :] = df_processed
                            error_logs.extend(end_errors)

                    # Check if the end coordinates actually contain valid data
                    if has_end_coords:
                        has_valid_end_coords = not df[end_lat_col].isna().all() and not df[end_lon_col].isna().all()
                        valid_pairs = ((df[start_lat_col].notna() & df[start_lon_col].notna()) & (df[end_lat_col].notna() & df[end_lon_col].notna())).any()
                    else:
                        has_valid_end_coords = False
                        valid_pairs = False
                    
                    # Determine geometry type based on available coordinates, actual data, and sheet type
                    # MARKA sheets should use MultiPoint geometry
                    if is_marka_sheet:
                        print(f"Processing '{sheet_name}' as MultiPoint (MARKA sheet)")

                        # First check if we have any valid coordinates before applying
                        has_valid_coords = ((df[start_lat_col].notna() & df[start_lon_col].notna()) | 
                                            (has_end_coords and df[end_lat_col].notna() & df[end_lon_col].notna())).any()
                        
                        if not has_valid_coords:
                            print(f"⚠️ Skipping '{sheet_name}' (No valid coordinates found in MARKA sheet)")
                            continue
                            
                        # Create MultiPoint geometry only for rows with valid coordinates
                        def create_multipoint(row):
                            try:
                                points = []
                                # Add start point if valid
                                if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]):
                                    points.append((float(row[start_lon_col]), float(row[start_lat_col])))
                                    
                                # Add end point if valid and available
                                if has_end_coords and pd.notna(row[end_lon_col]) and pd.notna(row[end_lat_col]):
                                    points.append((float(row[end_lon_col]), float(row[end_lat_col])))
                                    
                                # Return MultiPoint if we have points, else None
                                return MultiPoint(points) if points else None
                            except Exception as e:
                                print(f"Error creating MultiPoint for row {row.name}: {str(e)}")
                                return None
                        
                        # Apply the function to create geometry
                        df["geometry"] = df.apply(create_multipoint, axis=1)
                        
                        # Exclude coordinate columns from properties
                        exclude_cols = [start_lat_col, start_lon_col]
                        if has_end_coords:
                            exclude_cols.extend([end_lat_col, end_lon_col])
                        exclude_cols.append("geometry")
                        
                    # PAGAR PENGAMAN sheets with valid start/end coordinates should use LineString
                    elif is_pagar_pengaman_sheet and has_valid_end_coords and valid_pairs:
                        print(f"Processing '{sheet_name}' as LineString (PAGAR PENGAMAN sheet)")
                        
                        # Create LineString geometry with additional debugging and more flexible validation
                        def create_linestring(row):
                            try:
                                # Enhanced debugging
                                start_lat = row[start_lat_col]
                                start_lon = row[start_lon_col]
                                end_lat = row[end_lat_col]
                                end_lon = row[end_lon_col]
                                
                                # Skip processing if all coordinates are blank/NaN
                                if pd.isna(start_lat) and pd.isna(start_lon) and pd.isna(end_lat) and pd.isna(end_lon):
                                    return None
                                    
                                # First try to parse any string coordinates
                                if isinstance(start_lat, str):
                                    start_lat = parse_coordinate(start_lat)
                                if isinstance(start_lon, str):
                                    start_lon = parse_coordinate(start_lon)
                                if isinstance(end_lat, str):
                                    end_lat = parse_coordinate(end_lat)
                                if isinstance(end_lon, str):
                                    end_lon = parse_coordinate(end_lon)
                                
                                # Create LineString if we have valid coordinates
                                if (pd.notna(start_lon) and pd.notna(start_lat) and 
                                    pd.notna(end_lon) and pd.notna(end_lat)):
                                    return LineString([
                                        (float(start_lon), float(start_lat)),
                                        (float(end_lon), float(end_lat))
                                    ])
                                else:
                                    # If one set of coordinates is missing, fall back to Point
                                    if pd.notna(start_lon) and pd.notna(start_lat):
                                        offset = 0.0001
                                        return LineString([
                                            (float(start_lon), float(start_lat)),
                                            (float(start_lon) + offset, float(start_lat) + offset)
                                        ])
                                    return None
                            except Exception as e:
                                print(f"Error creating LineString for row {row.name}: {str(e)}")
                                return None
                        
                        df["geometry"] = df.apply(create_linestring, axis=1)
                        exclude_cols = [start_lat_col, start_lon_col, end_lat_col, end_lon_col, "geometry"]
                        
                    else:
                        # Other sheets use Point geometry (only start coordinates)
                        print(f"Processing '{sheet_name}' as Point geometry (regular sheet)")
                        
                        # Create Point geometry with start coordinates, only for rows with valid data
                        df["geometry"] = df.apply(
                            lambda row: Point(row[start_lon_col], row[start_lat_col]) 
                            if pd.notna(row[start_lon_col]) and pd.notna(row[start_lat_col]) else None,
                            axis=1
                        )
                        
                        # Exclude coordinate columns from properties
                        exclude_cols = [start_lat_col, start_lon_col]
                        if has_end_coords:
                            exclude_cols.extend([end_lat_col, end_lon_col])
                        exclude_cols.append("geometry")

                    # Drop rows where geometry is None
                    df = df.dropna(subset=["geometry"]).reset_index(drop=True)

                    # Only create GeoDataFrame if there are valid geometries
                    if len(df) > 0 and not df["geometry"].isnull().all():
                        # Filter REKAP columns before creating properties
                        properties_cols = [col for col in df.columns if col not in exclude_cols]
                        properties_cols = [col for col in properties_cols if 'rekap' not in str(col).lower()]
                        
                        # Create GeoDataFrame with explicit geometry setting
                        gdf = gpd.GeoDataFrame(df[properties_cols + ["geometry"]], geometry="geometry",crs="EPSG:4326")
                        
                        # Verify that the geometry column is properly set
                        if not gdf.geometry.name == "geometry":
                            gdf = gdf.set_geometry("geometry")
                        
                        gdf.columns = clean_column_names(gdf.columns)

                        gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
                        
                        gdf = gdf.loc[:, ~gdf.columns.astype(str).str.match(r"^None$|None_", na=False)]

                        gdf.columns = gdf.columns.astype(str).str.replace(r"\sNone\b", "", regex=True).str.strip()

                        # One final check for REKAP columns before saving
                        if any('rekap' in str(col).lower() for col in gdf.columns):
                            gdf = gdf.loc[:, ~gdf.columns.astype(str).str.contains("rekap", case=False, na=False)]
                        
                        # Define output file path
                        output_path = os.path.join(output_folder, f"{excel_name}_{sheet_name}.geojson")
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)

                        # Save as GeoJSON with additional parameters for image documentation
                        save_to_geojson(
                            gdf, 
                            output_path, 
                            batas_wilayah,
                            excel_name=excel_name,
                            sheet_name=sheet_name,
                            output_base_dir=output_base_dir
                        )
                    
                except Exception as e:
                    print(f"❌ Error processing sheet '{sheet_name}': {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue

            return error_logs
        except Exception as e:
            print(f"❌ Error processing file: {str(e)}")
            import traceback
            traceback.print_exc()
            return error_logs

    def process_single_excel_file(file_path, output_base_folder, batas_wilayah_path=None): # Process a single Excel file and convert it to GeoJSON
        # Create output folder
        output_folder = os.path.join(output_base_folder, "Extract GeoJSON")
        os.makedirs(output_folder, exist_ok=True)
        
        # Load the city boundaries shapefile if provided
        batas_wilayah = None
        if batas_wilayah_path and os.path.exists(batas_wilayah_path):
            try:
                batas_wilayah = gpd.read_file(batas_wilayah_path)
                print(f"✅ Loaded city boundaries from: {batas_wilayah_path}")
            except Exception as e:
                print(f"❌ Error loading city boundaries shapefile: {str(e)}")
        else:
            print("❌ No city boundaries provided or file not found.")
        
        # Get file name
        file_name = os.path.basename(file_path)
        excel_name = os.path.splitext(file_name)[0]

        error_logs = []
        try:
            # Process the file
            error_logs = flatten_excel_to_geojson(file_path, output_folder, excel_name, batas_wilayah, error_logs)
            print(f"✅ Completed processing: {file_name}")
            if error_logs:
                print(f"⚠️ Found {len(error_logs)} coordinate errors during processing")
        except Exception as e:
            print(f"❌ Error processing {file_name}: {str(e)}")
            import traceback
            traceback.print_exc()

        print(f"\nProcessing: {file_name}")
        print(f"\n🎉 Excel file processed. Output saved to: {output_folder}")

    def process_excel_folder_geojson(input_folder, output_base_folder, batas_wilayah_path=None): # Process a folder contain excel files and convert it to GeoJSON
        output_folder = os.path.join(output_base_folder, "Extract GeoJSON")
        os.makedirs(output_folder, exist_ok=True)
        
        # Load the city boundaries shapefile if provided
        batas_wilayah = None
        if batas_wilayah_path and os.path.exists(batas_wilayah_path):
            try:
                batas_wilayah = gpd.read_file(batas_wilayah_path)
                print(f"✅ Loaded city boundaries from: {batas_wilayah_path}")
            except Exception as e:
                print(f"❌ Error loading city boundaries shapefile: {str(e)}")
        else:
            print("❌ No city boundaries provided or file not found.")
        
        # Get all Excel files in the input folder
        excel_extensions = ['*.xlsx', '*.xls', '*.xlsm']
        excel_files = []
        
        for ext in excel_extensions:
            excel_files.extend(glob.glob(os.path.join(input_folder, ext)))
        
        # Initialize error log list
        all_error_logs = []
        
        # Process each Excel file
        for i, file_path in enumerate(excel_files, 1):
            file_name = os.path.basename(file_path)
            excel_name = os.path.splitext(file_name)[0]
            print(f"\n[{i}/{len(excel_files)}] Processing: {file_name}")
            
            try:
                # Collect errors from processing this file
                file_errors = flatten_excel_to_geojson(file_path, output_folder, excel_name, batas_wilayah, [])
                all_error_logs.extend(file_errors)
                print(f"✅ Completed processing: {file_name}")
                if file_errors:
                    print(f"⚠️ Found {len(file_errors)} coordinate errors during processing")
            except Exception as e:
                print(f"❌ Error processing {file_name}: {str(e)}")
            
        print(f"\n🎉 All Excel files processed. Output saved to: {output_folder}")