# %% [markdown]
# # Automation Excel to GeoJSON and Shapefile

# %% [markdown]
# ## 1. Application to Export Excel into GeoJson

# %% [markdown]
# ### 1.1. Function Codes

# %%
import sys
import subprocess
import pkg_resources
import os
import re
import io
import glob
import json
from pathlib import Path

def install_requirements():
    """Install required packages if they are not already installed."""
    required = {
        'pandas': '1.0.0',
        'openpyxl': '3.0.0',
        'geopandas': '0.9.0',
        'shapely': '1.7.0',
        'openpyxl_image_loader': '1.0.0',
    }
    
    # Check what's installed
    installed = {pkg.key: pkg.version for pkg in pkg_resources.working_set}
    
    # Determine what needs to be installed
    missing = []
    update = []
    
    for package, min_version in required.items():
        if package not in installed:
            missing.append(package)
        elif pkg_resources.parse_version(installed[package]) < pkg_resources.parse_version(min_version):
            update.append(package)
    
    # If packages need to be installed or updated
    if missing or update:
        print("Some required packages are missing or need to be updated.")
        print(f"Missing: {', '.join(missing) if missing else 'None'}")
        print(f"Need update: {', '.join(update) if update else 'None'}")
        
        try:
            # Install missing packages
            if missing:
                print(f"Installing missing packages: {', '.join(missing)}")
                subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
            
            # Update packages that need updating
            if update:
                print(f"Updating packages: {', '.join(update)}")
                subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade"] + update)
                
            print("All required packages have been installed/updated successfully!")
            
        except Exception as e:
            print(f"Failed to install required packages: {str(e)}")
            print("Please manually install the required packages using:")
            print("pip install pandas openpyxl geopandas shapely openpyxl_image_loader")
            sys.exit(1)

install_requirements()

# Import required libraries
import pandas as pd
from openpyxl import load_workbook
import geopandas as gpd
from shapely.geometry import Point, LineString, MultiPoint
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter, column_index_from_string

# Global dictionary to store image paths for each Excel file, sheet, and row
image_paths_registry = {}

def extract_images_from_excel(file_path, output_folder):
    """Extract images from columns containing specific keywords in all sheets and save them."""
    # Keywords to look for in column names
    KEYWORDS = ["DOKUMENTASI", "RAMBU", "RPPJ"]
    
    # Create subdirectories for each category directly in the main output folder
    dokumentasi_folder = os.path.join(output_folder, "Dokumentasi")
    rambu_folder = os.path.join(output_folder, "Rambu")
    rppj_folder = os.path.join(output_folder, "RPPJ")
    
    for folder in [dokumentasi_folder, rambu_folder, rppj_folder]:
        os.makedirs(folder, exist_ok=True)
    
    # Initialize image paths registry for this file
    file_name = os.path.basename(file_path)
    if file_name not in image_paths_registry:
        image_paths_registry[file_name] = {}
    
    try:
        # Extract filename from path
        file_name_clean = re.search(r'([^\\]+)\.xlsx$', file_path)
        if file_name_clean:
            file_name_clean = file_name_clean.group(1)
        else:
            file_name_clean = os.path.basename(file_path).replace('.xlsx', '')
        
        # Process each sheet independently to prevent file handle issues
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames.copy()  # Make a copy of sheet names before closing
        wb.close()  # Close immediately to avoid keeping file handles open
        
        successful_sheets = 0
        total_images_saved = 0
        
        # Process sheets in the order they appear in the workbook
        for sheet_idx, sheet_name in enumerate(sheet_names, 1):
            try:
                print(f"Processing sheet {sheet_idx}/{len(sheet_names)}: {sheet_name}")
                
                # Initialize image paths registry for this sheet
                if sheet_name not in image_paths_registry[file_name]:
                    image_paths_registry[file_name][sheet_name] = {}
                
                # Create a safe sheet name for filenames
                safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
                
                # Open a fresh workbook instance for each sheet
                wb = load_workbook(file_path, data_only=True)
                ws = wb[sheet_name]
                
                # Create a fresh image loader for this sheet
                image_loader = SheetImageLoader(ws)
                
                # Step 1: Extract merged column headers from rows 1-5
                # This creates a mapping from column index to full column name
                column_names = {}
                
                # Get values for rows 1-5 for each column
                for col in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col)
                    header_values = []
                    
                    for row in range(1, 6):  # Rows 1-5
                        cell_value = ws[f"{col_letter}{row}"].value
                        if cell_value:
                            header_values.append(str(cell_value).strip())
                    
                    # Combine the header parts into one name
                    if header_values:
                        column_names[col] = " ".join(header_values)
                    else:
                        column_names[col] = f"Column_{col_letter}"
                
                # Step 2: Find columns containing our keywords
                dokumentasi_columns = {}
                rambu_columns = {}
                rppj_columns = {}
                nama_rambu_column = None
                jenis_tiang_column = None
                
                # First, find all our target columns
                for col, name in column_names.items():
                    if "DOKUMENTASI" in name.upper():
                        dokumentasi_columns[col] = name
                    elif "RAMBU" in name.upper() and "NAMA RAMBU" not in name.upper():
                        rambu_columns[col] = name
                    elif "RPPJ" in name.upper():
                        rppj_columns[col] = name
                    
                    # Find the specific column for "Nama Rambu"
                    if "NAMA RAMBU" in name.upper():
                        nama_rambu_column = col
                        print(f"Found 'Nama Rambu' column: {name} (Column {get_column_letter(col)})")
                    
                    # Find the specific column for "Jenis Tiang"
                    elif "JENIS TIANG" in name.upper():
                        jenis_tiang_column = col
                        print(f"Found 'Jenis Tiang' column: {name} (Column {get_column_letter(col)})")
                
                # Track images processed for each category
                images_by_category = {
                    "dokumentasi": 0,
                    "rambu": 0,
                    "rppj": 0
                }
                
                # Process each category separately
                image_paths = {}  # Track image paths for this sheet
                
                # 1. Process DOKUMENTASI columns
                if dokumentasi_columns:
                    processed, paths = process_image_columns(
                        ws, image_loader, dokumentasi_columns, dokumentasi_folder, 
                        file_name_clean, safe_sheet_name, "dokumentasi", None, None
                    )
                    images_by_category["dokumentasi"] += processed
                    image_paths.update(paths)
                
                # 2. Process RAMBU columns (custom naming based on "Nama Rambu" column)
                if rambu_columns:
                    if nama_rambu_column is None:
                        print("⚠️ Warning: 'Nama Rambu' column not found. Using default naming for Rambu images.")
                    
                    processed, paths = process_image_columns(
                        ws, image_loader, rambu_columns, rambu_folder, 
                        file_name_clean, safe_sheet_name, "rambu", nama_rambu_column, None
                    )
                    images_by_category["rambu"] += processed
                    image_paths.update(paths)
                
                # 3. Process RPPJ columns (custom naming based on "Jenis Tiang" column)
                if rppj_columns:
                    if jenis_tiang_column is None:
                        print("⚠️ Warning: 'Jenis Tiang' column not found. Using default naming for RPPJ images.")
                    
                    processed, paths = process_image_columns(
                        ws, image_loader, rppj_columns, rppj_folder, 
                        file_name_clean, safe_sheet_name, "rppj", None, jenis_tiang_column
                    )
                    images_by_category["rppj"] += processed
                    image_paths.update(paths)
                
                # Store image paths for this sheet
                image_paths_registry[file_name][sheet_name] = image_paths
                
                total_images_saved += sum(images_by_category.values())
                
                print(f"✅ Completed sheet '{sheet_name}':")
                print(f"  - Dokumentasi: {images_by_category['dokumentasi']} images")
                print(f"  - Rambu: {images_by_category['rambu']} images")
                print(f"  - RPPJ: {images_by_category['rppj']} images")
                
                if sum(images_by_category.values()) > 0:
                    successful_sheets += 1
                
            except Exception as e:
                print(f"❌ Error processing sheet '{sheet_name}' in file '{file_name_clean}': {str(e)}")
            
            finally:
                # Always close the workbook after processing each sheet
                if 'wb' in locals() and wb is not None:
                    try:
                        wb.close()
                    except:
                        pass
        
        print(f"✅ Completed processing file: {file_name_clean}")
        print(f"  - {successful_sheets}/{len(sheet_names)} sheets processed")
        print(f"  - {total_images_saved} total images extracted")
        return True
        
    except Exception as e:
        # Handle any errors in the outer scope
        print(f"❌ Error processing file '{file_path}': {str(e)}")
        return False

def process_image_columns(ws, image_loader, target_columns, output_folder, file_name_clean, 
                          safe_sheet_name, category, nama_rambu_column, jenis_tiang_column):
    """Process images in specified columns with custom naming logic."""
    
    # Track existing image names (to avoid duplicates for Rambu)
    existing_images = {}
    
    # Track image paths for each row
    image_paths = {}
    
    # Find image cells in target columns
    image_cells = {}
    for col in target_columns:
        col_letter = get_column_letter(col)
        column_name = target_columns[col]
        
        # Make column name safe for filename
        safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
        
        # Scan rows starting from row 6
        for row in range(6, ws.max_row + 1):
            cell_address = f"{col_letter}{row}"
            if image_loader.image_in(cell_address):
                image_cells[(row, col)] = {
                    'cell_address': cell_address,
                    'column_name': safe_column_name,
                    'column_letter': col_letter,
                    'row_number': row
                }
    
    if not image_cells:
        print(f"⚠️ No images found in {category.upper()} columns, skipping...")
        return 0, {}
    
    print(f"Found {len(image_cells)} images in {category.upper()} columns")
    
    # Process images in column-first, row-second order
    successful_images = 0
    image_cells_by_column = {}
    
    # Group by column
    for (row, col), cell_info in image_cells.items():
        if col not in image_cells_by_column:
            image_cells_by_column[col] = []
        image_cells_by_column[col].append((row, cell_info))
    
    # Process each column
    for col in sorted(image_cells_by_column.keys()):
        column_letter = get_column_letter(col)
        column_name = target_columns[col]
        safe_column_name = re.sub(r'[\\/*?:"<>|]', "_", column_name)
        
        print(f"Processing {len(image_cells_by_column[col])} images in column '{column_name}'")
        
        # Process rows in order
        for row, cell_info in sorted(image_cells_by_column[col], key=lambda x: x[0]):
            cell_address = cell_info['cell_address']
            try:
                # Get the image
                img = image_loader.get(cell_address)
                
                # Generate filename based on category
                if category == "rambu" and nama_rambu_column is not None:
                    # For RAMBU: Use "Nama Rambu" column value as filename
                    nama_rambu_value = ws[f"{get_column_letter(nama_rambu_column)}{row}"].value
                    
                    if nama_rambu_value and str(nama_rambu_value).strip():
                        # Use the actual value from "Nama Rambu" column
                        safe_nama_rambu = re.sub(r'[\\/*?:"<>|]', "_", str(nama_rambu_value).strip())
                        img_filename = f"{safe_nama_rambu}.png"
                        
                        # Check if this name already exists (to avoid duplicates)
                        if img_filename in existing_images:
                            print(f"  ⚠️ Duplicate 'Nama Rambu' found: {safe_nama_rambu} - Replacing existing image")
                        
                        existing_images[img_filename] = True
                    else:
                        # Fallback to default naming if no nama_rambu value
                        print(f"  ⚠️ No 'Nama Rambu' value found for row {row}, using default naming")
                        row_identifier = f"Row{row}"
                        safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                        img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                
                elif category == "rppj" and jenis_tiang_column is not None:
                    # For RPPJ: Use "Jenis Tiang" column value in filename
                    jenis_tiang_value = ws[f"{get_column_letter(jenis_tiang_column)}{row}"].value
                    safe_jenis_tiang = "Unknown"
                    if jenis_tiang_value and str(jenis_tiang_value).strip():
                        safe_jenis_tiang = re.sub(r'[\\/*?:"<>|]', "_", str(jenis_tiang_value).strip())
                    
                    # Get row identifier - ALWAYS use just the row number without column A value
                    row_identifier = f"Row{row}"
                    
                    safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                    img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_jenis_tiang}_{safe_row_identifier}.png"
                
                else:
                    # Default naming for DOKUMENTASI (and fallback for others)
                    # ALWAYS use just the row number without column A value
                    row_identifier = f"Row{row}"
                    
                    safe_row_identifier = re.sub(r'[\\/*?:"<>|]', "_", row_identifier)
                    img_filename = f"{file_name_clean}_Sheet_{safe_sheet_name}_Column_{safe_column_name}_{safe_row_identifier}.png"
                
                # Save the image
                img_path = os.path.join(output_folder, img_filename)
                with io.BytesIO() as img_buffer:
                    img.save(img_buffer, format="PNG")
                    img_buffer.seek(0)
                    with open(img_path, 'wb') as f:
                        f.write(img_buffer.read())
                
                # Store image path for this row
                if row not in image_paths:
                    image_paths[row] = []
                image_paths[row].append({
                    'category': category,
                    'column_name': column_name,
                    'path': img_path
                })
                
                successful_images += 1
                print(f"  ✅ Saved: {img_filename}")
            except Exception as e:
                print(f"  ❌ Error saving image at {cell_address}: {str(e)}")
    
    return successful_images, image_paths

def unique_column_names(columns):
    """Ensure column names are unique by appending suffix."""
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def clean_column_names(columns):
    """Standardize column names by capitalizing each word properly."""
    cleaned_columns = []
    seen = {}
    for col in columns:
        col = str(col).strip()
        
        # Remove columns with "Rekap" in the name
        if "rekap" in col.lower():
            continue
        
        # Fix for duplicated words (like "No No", "Detail Lokasi Detail Lokasi")
        words = col.split()
        if len(words) >= 2:
            # Check for repeated word patterns
            half_length = len(words) // 2
            if words[:half_length] == words[half_length:] and len(words) % 2 == 0:
                # If the first half matches the second half, only use the first half
                col = " ".join(words[:half_length])
        
        # Capitalize words
        col = " ".join(word.capitalize() for word in col.split())
        
        # Handle duplicates
        if col in seen:
            seen[col] += 1
            col = f"{col} {seen[col]}"
        else:
            seen[col] = 0
        
        cleaned_columns.append(col)
    return cleaned_columns

def fix_coordinates(row, lat_col, lon_col):
    """Fix latitude and longitude values that may be in the wrong format."""
    lat, lon = row[lat_col], row[lon_col]
    original_lat, original_lon = lat, lon  # Store original values for debugging
    
    # Handle string values with commas (e.g., "-69,694,951" or "1,065,663,308")
    if isinstance(lat, str):
        try:
            lat = float(lat.replace(',', ''))
        except (ValueError, AttributeError):
            lat = pd.NA
    
    if isinstance(lon, str):
        try:
            lon = float(lon.replace(',', ''))
        except (ValueError, AttributeError):
            lon = pd.NA
    
    # First attempt to detect the scale by number of digits
    if pd.notna(lat):
        lat_abs = abs(lat)
        # For values like -6448977 (which should be around -6.4 degrees)
        if 1_000_000 < lat_abs < 10_000_000 and str(int(lat_abs)).startswith(('6', '7', '8', '9')):
            lat = lat / 1_000_000
        elif lat_abs > 90:
            # General scaling rules based on magnitude
            if lat_abs > 10_000_000:  # Very large values
                lat = lat / 10_000_000
            elif lat_abs > 1_000_000:  # Large values (common for Indonesia coords)
                lat = lat / 1_000_000
            elif lat_abs > 90_000:
                lat = lat / 1_000
    
    if pd.notna(lon):
        lon_abs = abs(lon)
        # For Indonesian longitudes (usually around 106-110 degrees)
        if 100_000_000 < lon_abs < 1_500_000_000:
            lon = lon / 10_000_000
        elif lon_abs > 180:
            # General scaling rules based on magnitude
            if lon_abs > 10_000_000:
                lon = lon / 10_000_000
            elif lon_abs > 1_000_000:
                lon = lon / 1_000_000
            elif lon_abs > 180_000:
                lon = lon / 1_000
    
    # Final validation with better error reporting
    if pd.notna(lat) and (lat < -90 or lat > 90):
        print(f"Warning: Invalid latitude after correction: {lat} (original: {original_lat})")
        
        # Last attempt for specific problematic values
        if -10000 < lat < -90 or 90 < lat < 10000:
            lat = lat / 100
            print(f"  Attempting additional scaling: now {lat}")
        
        if lat < -90 or lat > 90:  # Still invalid
            lat = pd.NA
    
    if pd.notna(lon) and (lon < -180 or lon > 180):
        print(f"Warning: Invalid longitude after correction: {lon} (original: {original_lon})")
        
        # Last attempt for specific problematic values
        if -18000 < lon < -180 or 180 < lon < 18000:
            lon = lon / 100
            print(f"  Attempting additional scaling: now {lon}")
            
        if lon < -180 or lon > 180:  # Still invalid
            lon = pd.NA
        
    return pd.Series([lat, lon])

def clean_geojson(gdf, output_path):
    """Save GeoDataFrame in a clean format GeoJSON file and delete the temp file."""
    temp_path = output_path.replace(".geojson", "_temp.geojson")
    gdf.to_file(temp_path, driver="GeoJSON")

    with open(temp_path, "r", encoding="utf-8") as file:
        geojson_data = json.load(file)

    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(geojson_data, file, indent=4)
    
    # Delete the temporary file
    try:
        os.remove(temp_path)
        print(f"✅ Deleted temporary file: {temp_path}")
    except Exception as e:
        print(f"⚠️ Could not delete temporary file {temp_path}: {str(e)}")

    print(f"✅ Saved: {output_path}")

def find_coordinate_columns(df, prefix, column_type):
    """Find coordinate columns with various naming patterns."""
    patterns = []
    
    if prefix == 'start':
        if column_type == 'lat':
            patterns = [
                'awal latitude', 'awal lat', 
                'koordinat latitude', 'koordinat lat',
                'latitude awal', 'lat awal',
                'latitude 1', 'lat 1'
            ]
        else:  # longitude
            patterns = [
                'awal longitude', 'awal lon', 
                'koordinat longitude', 'koordinat lon',
                'longitude awal', 'lon awal',
                'longitude 1', 'lon 1'
            ]
    else:  # end
        if column_type == 'lat':
            patterns = [
                'akhir latitude', 'akhir lat',
                'latitude akhir', 'lat akhir',
                'latitude 2', 'lat 2'
            ]
        else:  # longitude
            patterns = [
                'akhir longitude', 'akhir lon',
                'longitude akhir', 'lon akhir',
                'longitude 2', 'lon 2'
            ]
    
    # If we're looking for start coordinates, and just a single coordinate exists (no start/end distinction)
    if prefix == 'start' and column_type == 'lat':
        # Add cases where latitude exists without start/end specifier
        patterns.extend(['latitude', 'lat'])
    elif prefix == 'start' and column_type == 'lon':
        # Add cases where longitude exists without start/end specifier
        patterns.extend(['longitude', 'lon'])
    
    # Search for column containing any of the patterns
    for pattern in patterns:
        for col in df.columns:
            if isinstance(col, str) and pattern in col.lower():
                return col
    
    return None

def flatten_excel_to_geojson(file_path, output_folder, images_folder):
    """Convert all sheets from an Excel file to GeoJSON with image paths."""
    file_name = os.path.basename(file_path)
    file_basename_no_ext = os.path.splitext(file_name)[0]
    
    # Load workbook
    wb = load_workbook(file_path, data_only=True)
    
    # Keep track of which rows were kept for each sheet
    row_mapping = {}

    for sheet_name in wb.sheetnames:
        sheet_row_mapping = {}  # Store mapping from final index to original Excel row
        
        ws = wb[sheet_name]
        
        # Print sheet dimensions for debugging
        print(f"Processing sheet: {sheet_name} (Rows: {ws.max_row}, Columns: {ws.max_column})")

        # Unmerge cells and fill values
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(str(merge))
            top_left = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    ws.cell(row, col, top_left)

        # Convert to DataFrame
        data = list(ws.values)
        df = pd.DataFrame(data)

        # Check if DataFrame is empty or has too few rows
        if df.empty or len(df) < 3:
            print(f"⚠️ Skipping '{sheet_name}' (Empty sheet or insufficient data)")
            continue

        # Find rows containing "NO" (safer approach)
        try:
            header_indices = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index
            if len(header_indices) == 0:
                print(f"⚠️ Skipping '{sheet_name}' (No header row with 'NO' found)")
                continue
            header_index = header_indices[0]
        except Exception as e:
            print(f"⚠️ Skipping '{sheet_name}' (Error finding header row: {str(e)})")
            continue

        # Use the identified header row
        df.columns = df.iloc[header_index].astype(str).str.strip()

        # Remove empty columns
        df = df.dropna(axis=1, how="all")

        # First pass of REKAP filtering
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns - {str(e)}")

        # Drop rows above and including header, plus the empty row after header
        rows_to_drop = list(range(0, header_index + 2))
        if len(df) > max(rows_to_drop) + 1:
            # Track which rows were kept and their original Excel row numbers
            kept_rows = [i for i in range(len(df)) if i not in rows_to_drop]
            for new_idx, old_idx in enumerate(kept_rows):
                # Original Excel row number is old_idx + 1 (since Excel is 1-indexed)
                sheet_row_mapping[new_idx] = old_idx + 1
                
            df = df.drop(index=rows_to_drop).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough data rows after header)")
            continue

        # Improved header row merging
        if len(df) >= 2:
            # Convert values to strings
            first_row = df.iloc[0].astype(str).replace('None', '').replace('nan', '')
            second_row = df.iloc[1].astype(str).replace('None', '').replace('nan', '')
            
            # Smart merging
            merged_header = []
            for a, b in zip(first_row, second_row):
                a = a.strip()
                b = b.strip()
                
                if "rekap" in a.lower() or "rekap" in b.lower():
                    merged_header.append("TO_BE_REMOVED")
                    continue
                
                if not a and not b:
                    merged_header.append("Column_" + str(len(merged_header)))
                elif not a:
                    merged_header.append(b)
                elif not b:
                    merged_header.append(a)
                else:
                    if a.lower() in b.lower():
                        merged_header.append(b)
                    elif b.lower() in a.lower():
                        merged_header.append(a)
                    else:
                        merged_header.append(f"{a} {b}")
            
            # Ensure column names are unique
            df.columns = unique_column_names(merged_header)
            
            # Remove columns marked for removal
            df = df.loc[:, ~df.columns.str.contains("TO_BE_REMOVED")]
            
            # Update row mapping to account for removed header rows
            new_sheet_row_mapping = {}
            for new_idx, old_idx in enumerate(range(2, len(df))):
                if old_idx in sheet_row_mapping:
                    new_sheet_row_mapping[new_idx] = sheet_row_mapping[old_idx]
            sheet_row_mapping = new_sheet_row_mapping
            
            # Remove the first two rows used for headers
            df = df.drop(index=[0, 1]).reset_index(drop=True)
        else:
            print(f"⚠️ Skipping '{sheet_name}' (Not enough rows for headers)")
            continue

        # Store the final row mapping for this sheet
        row_mapping[sheet_name] = sheet_row_mapping
        
        # Print the row mapping for debugging
        print(f"Row mapping for sheet {sheet_name}:")
        for new_idx, excel_row in sheet_row_mapping.items():
            print(f"  GeoJSON idx {new_idx} -> Excel row {excel_row}")

        # Second pass of REKAP filtering
        try:
            rekap_mask = df.columns.astype(str).str.contains("REKAP", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns (second pass) - {str(e)}")

        # Normalize column names
        df.columns = [str(col).lower().strip() if col is not None else f"col_{i}" 
                      for i, col in enumerate(df.columns)]
        
        # Third pass of REKAP filtering
        try:
            rekap_mask = df.columns.str.contains("rekap", case=False, na=False)
            if rekap_mask.any():
                df = df.loc[:, ~rekap_mask]
        except Exception as e:
            print(f"Warning in '{sheet_name}': Error filtering REKAP columns (third pass) - {str(e)}")

        # Find coordinate columns
        start_lat_col = find_coordinate_columns(df, 'start', 'lat')
        start_lon_col = find_coordinate_columns(df, 'start', 'lon')
        end_lat_col = find_coordinate_columns(df, 'end', 'lat')
        end_lon_col = find_coordinate_columns(df, 'end', 'lon')
        
        # Check if lat/lon columns found
        if not start_lat_col or not start_lon_col:
            print(f"⚠️ Skipping '{sheet_name}' (No latitude/longitude columns found)")
            continue
        
        # Fix coordinates
        df[[start_lat_col, start_lon_col]] = df.apply(
            lambda row: fix_coordinates(row, start_lat_col, start_lon_col), axis=1
        )
        
        if end_lat_col and end_lon_col:
            df[[end_lat_col, end_lon_col]] = df.apply(
                lambda row: fix_coordinates(row, end_lat_col, end_lon_col), axis=1
            )
        
        # Drop rows with invalid coordinates and update row mapping
        valid_rows = df.dropna(subset=[start_lat_col, start_lon_col]).index
        updated_mapping = {new_idx: sheet_row_mapping[old_idx] for new_idx, old_idx in enumerate(valid_rows) if old_idx in sheet_row_mapping}
        sheet_row_mapping = updated_mapping
        df = df.dropna(subset=[start_lat_col, start_lon_col]).reset_index(drop=True)
        
        if df.empty:
            print(f"⚠️ Skipping '{sheet_name}' (No valid coordinates found)")
            continue
        
        # Add original Excel row numbers to the dataframe for debugging
        df['original_excel_row'] = df.index.map(lambda idx: sheet_row_mapping.get(idx, 'unknown'))
        
        # Update row mapping for the sheet
        row_mapping[sheet_name] = sheet_row_mapping
        
        # Add image paths to dataframe
        df['image_paths'] = df.index.map(lambda idx: [])
        
        # Create GeoDataFrame
        geometry = []
        for idx, row in df.iterrows():
            try:
                start_lat = row[start_lat_col]
                start_lon = row[start_lon_col]
                
                # Create start point
                start_point = Point(start_lon, start_lat)
                
                # If we have end coordinates, create a LineString
                if end_lat_col and end_lon_col and pd.notna(row[end_lat_col]) and pd.notna(row[end_lon_col]):
                    end_lat = row[end_lat_col]
                    end_lon = row[end_lon_col]
                    geometry.append(LineString([(start_lon, start_lat), (end_lon, end_lat)]))
                else:
                    # Otherwise, just use the start point
                    geometry.append(start_point)
            except Exception as e:
                print(f"Error creating geometry for row {idx}: {str(e)}")
                # Add a placeholder point to maintain alignment
                geometry.append(Point(0, 0))
        
        # Create GeoDataFrame
        gdf = gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")
        
        # Add Dokumentasi image paths using the correct row mapping
        file_basename = os.path.basename(file_path)
        file_basename_no_ext = os.path.splitext(file_basename)[0]
        
        # Now when we look for images, we'll use the original Excel row numbers
        gdf['Dokumentasi_Image_Paths'] = gdf.index.map(
            lambda idx: find_images_for_row(idx, sheet_row_mapping.get(idx, None), images_folder, 
                                           file_basename, file_basename_no_ext, sheet_name)
        )
        
        # Clean up columns
        gdf.columns = clean_column_names(gdf.columns)
        
        # Check if we have any valid geometries
        if len(gdf) == 0:
            print(f"⚠️ Skipping '{sheet_name}' (No valid geometries created)")
            continue
        
        # Save GeoJSON
        safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
        safe_file_name = re.sub(r'[\\/*?:"<>|]', "_", os.path.splitext(os.path.basename(file_path))[0])
        output_file = os.path.join(output_folder, f"{safe_file_name}_{safe_sheet_name}.geojson")
        
        # Save as GeoJSON
        clean_geojson(gdf, output_file)
        
    print(f"✅ Completed processing file: {os.path.basename(file_path)}")
    wb.close()
    return True

def find_images_for_row(idx, excel_row, images_folder, file_basename, file_basename_no_ext, sheet_name):
    """
    Find images for a specific row using the original Excel row number.
    """
    paths = []
    
    # If we have a valid Excel row number from our mapping
    if excel_row:
        # First check the image registry
        if file_basename in image_paths_registry and sheet_name in image_paths_registry[file_basename]:
            sheet_image_paths = image_paths_registry[file_basename][sheet_name]
            
            # The key insight: use the original Excel row number to look up images
            if excel_row in sheet_image_paths:
                for img_info in sheet_image_paths[excel_row]:
                    if img_info['category'].lower() == 'dokumentasi':
                        original_path = img_info['path']
                        filename = os.path.basename(original_path)
                        relative_path = f"../Extract Images/{file_basename_no_ext}/Dokumentasi/{filename}"
                        paths.append(relative_path)
                        print(f"Found match: GeoJSON idx {idx} -> Excel row {excel_row} -> {filename}")
    
    # If no images found through the registry, use the file system approach
    if not paths:
        dokumentasi_dir = os.path.join(images_folder, file_basename_no_ext, "Dokumentasi")
        if os.path.exists(dokumentasi_dir):
            # Search for files matching the Excel row number
            if excel_row:
                row_patterns = [f"Row{excel_row}", f"row{excel_row}", f"_Row{excel_row}", f"_row{excel_row}"]
                
                for filename in os.listdir(dokumentasi_dir):
                    for pattern in row_patterns:
                        if pattern in filename:
                            relative_path = f"../Extract Images/{file_basename_no_ext}/Dokumentasi/{filename}"
                            if relative_path not in paths:  # Avoid duplicates
                                paths.append(relative_path)
                                print(f"Found file match: GeoJSON idx {idx} -> Excel row {excel_row} -> {filename}")
    
    return paths

def main():
    """Main function to process Excel files with automatic folder structure."""
    print("Excel to GeoJSON with Image Extraction Utility")
    print("=============================================")
    
    # Get input folder from user
    input_folder = input("Enter the path to the folder containing Excel files: ").strip()
    
    # Validate input folder
    if not os.path.isdir(input_folder):
        print(f"❌ Error: The specified input folder '{input_folder}' does not exist.")
        return
    
    # Create main output folder structure automatically
    output_base_folder = os.path.join(os.path.dirname(input_folder), "Conversion Result")
    images_folder = os.path.join(output_base_folder, "Extract Images")
    geojson_folder = os.path.join(output_base_folder, "Convert GeoJSON")
    
    # Create output folders if they don't exist
    try:
        os.makedirs(output_base_folder, exist_ok=True)
        os.makedirs(images_folder, exist_ok=True)
        os.makedirs(geojson_folder, exist_ok=True)
        print(f"✅ Created output folder structure in: {output_base_folder}")
    except Exception as e:
        print(f"❌ Error: Could not create output folders: {str(e)}")
        return
    
    # Find all Excel files in input folder
    excel_files = glob.glob(os.path.join(input_folder, "*.xlsx"))
    
    if not excel_files:
        print("❌ No Excel files found in the input folder.")
        print(f"Please place Excel files in: {input_folder}")
        return
    
    print(f"Found {len(excel_files)} Excel files to process.")
    
    # Process each Excel file
    for i, file_path in enumerate(excel_files, 1):
        print(f"\n[{i}/{len(excel_files)}] Processing file: {os.path.basename(file_path)}")
        
        # Step 1: Extract images from Excel to the images folder directly
        print("\n=== STEP 1: Extracting images from Excel ===")
        # Change this line to use images_folder directly instead of file_images_folder
        images_extracted = extract_images_from_excel(file_path, images_folder)
        
        if not images_extracted:
            print(f"❌ Failed to extract images from {os.path.basename(file_path)}")
            continue
        
        # Step 2: Convert Excel to GeoJSON with image paths and save to GeoJSON folder
        print("\n=== STEP 2: Converting Excel to GeoJSON with image paths ===")
        geojson_converted = flatten_excel_to_geojson(file_path, geojson_folder, images_folder)
        
        if not geojson_converted:
            print(f"❌ Failed to convert {os.path.basename(file_path)} to GeoJSON")
            continue
        
        print(f"✅ Successfully processed {os.path.basename(file_path)}")
    
    print("\n=== PROCESSING COMPLETE ===")
    print(f"Images are in: {images_folder}")
    print(f"GeoJSON files are in: {geojson_folder}")

if __name__ == "__main__":
    main()


