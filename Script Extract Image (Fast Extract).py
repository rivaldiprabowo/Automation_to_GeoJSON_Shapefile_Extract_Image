# %% [markdown]
# # Application to Extract Image from Excel

# %% [markdown]
# ## 1. Import Library

# %%
import os
import pandas as pd
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image
import io
import re
import gc
from concurrent.futures import ThreadPoolExecutor
import time
from io import BytesIO
from openpyxl.utils import column_index_from_string, get_column_letter
from lxml import etree
import zipfile

# %% [markdown]
# ## 2. Application to Extract Images from Excel

# %% [markdown]
# ### 2.1. Function Codes

# %%
def unique_column_names(columns): 
    """Ensure column names are unique by appending a suffix."""
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def clean_column_names(columns): #Standardize column names by capitalizing each word properly.
    cleaned_columns = []
    seen = {}
    for col in columns:
        col = str(col).strip()
        col = " ".join(word.capitalize() for word in col.split())
        if col in seen:
            seen[col] += 1
            col = f"{col} {seen[col]}"
        else:
            seen[col] = 0
        cleaned_columns.append(col)
    return cleaned_columns

def extract_images_from_excel(file_path, output_folder):
    """Extract images from columns containing 'DOKUMENTASI' in all sheets and save them."""    
    try:
        # Extract filename from path
        file_name_clean = re.search(r'([^\\]+)\.xlsx$', file_path)
        if file_name_clean:
            file_name_clean = file_name_clean.group(1)
        else:
            file_name_clean = os.path.basename(file_path).replace('.xlsx', '')
        
        # Load workbook once for the entire process
        wb = load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Create a fresh image loader for each sheet
            image_loader = SheetImageLoader(ws)
            
            # Unmerge cells and fill values
            for merge in list(ws.merged_cells):
                ws.unmerge_cells(str(merge))
                top_left = ws.cell(merge.min_row, merge.min_col).value
                for row in range(merge.min_row, merge.max_row + 1):
                    for col in range(merge.min_col, merge.max_col + 1):
                        ws.cell(row, col, top_left)
            
            # Convert to DataFrame
            data = list(ws.values)
            df = pd.DataFrame(data)
            
            # Identify the header row
            try:
                header_index = df[df.apply(lambda x: x.astype(str).str.contains("NO", case=False, na=False)).any(axis=1)].index[0]
            except IndexError:
                print(f"⚠️ Could not identify header row in sheet '{sheet_name}' of file '{file_name_clean}', skipping...")
                continue
            
            # Set the header
            df.columns = df.iloc[2].astype(str).str.strip()
            
            # Remove empty columns
            df = df.dropna(axis=1, how="all")
            
            # Drop "REKAP" section if present
            df = df.loc[:, ~df.columns.str.contains("REKAP", case=False, na=False)]
            df = df.drop(index=[0, 1, 4]).reset_index(drop=True)
            
            # Merge first two rows if needed
            merged_header = [a if a == b else f"{a} {b}" for a, b in zip(df.iloc[0], df.iloc[1])]
            
            # Ensure column names are unique
            df.columns = unique_column_names(merged_header)
            
            # Remove the first two rows used for headers
            df = df.drop(index=[0, 1]).reset_index(drop=True)
            
            # Normalize column names for consistent detection
            df.columns = df.columns.str.upper().str.strip()

            # Apply column renaming after creating the GeoDataFrame
            df.columns = clean_column_names(df.columns)

            # Remove unwanted "None_" and "None" columns
            df = df.loc[:, ~df.columns.str.match(r"^None$|None_", na=False)]

            # Remove " None" from remaining column names
            df.columns = df.columns.str.replace(r"\sNone\b", "", regex=True).str.strip()
            
            # Find all columns that contain "DOKUMENTASI" (case-insensitive)
            dokumentasi_cols = [col for col in df.columns if "DOKUMENTASI" in col.upper()]
            
            # Skip processing if no dokumentasi columns are found
            if not dokumentasi_cols:
                print(f"⚠️ No 'DOKUMENTASI' columns found in sheet '{sheet_name}' of file '{file_name_clean}', skipping...")
                continue
            
            print(f"Found {len(dokumentasi_cols)} 'DOKUMENTASI' columns in sheet '{sheet_name}'")
            
            # First pass: collect all cell coordinates with images
            image_cells = []
            for col_idx, col_name in enumerate(df.columns):
                if col_name not in dokumentasi_cols:
                    continue
                
                # Calculate proper column index (Excel columns are 1-indexed)
                excel_col_idx = col_idx + 1
                
                for row_idx in range(2, ws.max_row + 1):
                    cell_address = ws.cell(row=row_idx, column=excel_col_idx).coordinate
                    if image_loader.image_in(cell_address):
                        image_cells.append((row_idx, excel_col_idx, col_name, cell_address))
            
            # Second pass: extract and save all images at once
            for row_idx, excel_col_idx, col_name, cell_address in image_cells:
                try:
                    # Get the image
                    img = image_loader.get(cell_address)
                    
                    # Make column name safe for filename
                    safe_col_name = col_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                    
                    # Save path for the image
                    img_path = os.path.join(output_folder, f"{file_name_clean}_Sheet_{sheet_name}_Column_{safe_col_name}_Row_{row_idx}.png")
                    
                    # Save the image directly to the file system
                    # Use a temporary BytesIO buffer to avoid file handle issues
                    with io.BytesIO() as img_buffer:
                        img.save(img_buffer, format="PNG")
                        img_buffer.seek(0)
                        with open(img_path, 'wb') as f:
                            f.write(img_buffer.read())
                    
                    print(f"✅ Image saved: {img_path}")
                except Exception as e:
                    print(f"❌ Error saving image at {cell_address} in file '{file_name_clean}': {str(e)}")
                    # Continue with next image instead of failing the entire process
        
        # Close the workbook after all processing is complete
        wb.close()
        
        print(f"✅ Completed processing file: {file_name_clean}")
        return True
    except Exception as e:
        # Ensure workbook is closed even if an error occurs
        try:
            if 'wb' in locals():
                wb.close()
        except:
            pass
        
        print(f"❌ Error processing file '{file_path}': {str(e)}")
        return False

def process_excel_folder(folder_path, export_folder):
    """Process all Excel files in a folder and extract images from them."""
    import os
    
    # Create a single "Extract Images" folder within the export directory
    output_folder = os.path.join(export_folder, "Extract Images")
    os.makedirs(output_folder, exist_ok=True)
    
    # Track statistics
    total_files = 0
    successful_files = 0
    failed_files = []
    
    # Get all Excel files in the folder - with full paths
    excel_files = []
    for file in os.listdir(folder_path):
        if file.endswith(('.xlsx', '.xlsm')):
            excel_files.append(os.path.join(folder_path, file))
    
    if not excel_files:
        print("⚠️ No Excel files found in the specified folder.")
        return
    
    total_files = len(excel_files)
    print(f"🔍 Found {total_files} Excel files to process.")
    print(f"🗂️ All images will be saved to: {output_folder}")
    
    # Process each Excel file
    for i, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print(f"\n📊 Processing file {i}/{total_files}: {file_name}")
        
        if extract_images_from_excel(file_path, output_folder):
            successful_files += 1
        else:
            failed_files.append(file_name)
    
    # Print summary
    print("\n" + "="*50)
    print("📈 PROCESSING SUMMARY")
    print("="*50)
    print(f"Total files: {total_files}")
    print(f"Successfully processed: {successful_files}")
    print(f"Failed to process: {len(failed_files)}")
    print(f"Images saved to: {output_folder}")
    
    if failed_files:
        print("\nFiles that could not be processed:")
        for file in failed_files:
            print(f"- {file}")
    
    print("\n🎉 All Excel files processing completed!")

# %% [markdown]
# ### 2.2. Run Function

# %%
excel_folder = r"C:\Users\kanzi\Documents\Part Time Job\Data Hasil Survey"  # Path to Excel files
export_folder = r"C:\Users\kanzi\Documents\Part Time Job\Hasil Export Images"  # Path for export results
        
# Run the function with your paths
process_excel_folder(excel_folder, export_folder)


